# ⚠️  WARNING — DESTRUCTIVE SCRIPT ⚠️
# This script permanently DELETES sheets from the shared team Excel file.
# It was used once for a one-off cleanup and should NEVER run again without
# careful review.  Running it would erase historical data for every designer
# whose sheet is not in the 'keep' list below.
#
# To recover lost data, use OneDrive version history on _TeamProduction.xlsx:
#   Right-click the file in the OneDrive web interface → Version history.
#
# Triple-guard:
#   1. Hard exit on import / direct run unless PRODCALC_DESTRUCTIVE_DIAG=1.
#   2. Explicit "--i-know-this-deletes-data" CLI flag required.
#   3. Interactive y/N confirmation (will hang in non-tty).

import os
import sys

if os.environ.get("PRODCALC_DESTRUCTIVE_DIAG") != "1":
    print("⚠️  Refusing to run destructive cleanup script.")
    print("    Set env var PRODCALC_DESTRUCTIVE_DIAG=1 AND pass")
    print("    --i-know-this-deletes-data on the command line if you really")
    print("    mean to delete every team sheet except Dashboard + Gerardo.")
    sys.exit(1)

if "--i-know-this-deletes-data" not in sys.argv:
    print("⚠️  Missing --i-know-this-deletes-data flag. Refusing to run.")
    sys.exit(1)

confirm = input("Type 'DELETE TEAM DATA' to proceed: ").strip()
if confirm != "DELETE TEAM DATA":
    print("Aborted.")
    sys.exit(1)

import openpyxl

team = r'C:\Users\gerar\Envista\SPARK-GLB-OPS-ICON - Reports\Productions\_TeamProduction.xlsx'
wb = openpyxl.load_workbook(team)
print("Before:", wb.sheetnames)

# Keep only Dashboard and Gerardo
to_remove = [s for s in wb.sheetnames if s not in ('Dashboard', 'Gerardo')]
for s in to_remove:
    del wb[s]
    print(f"Removed sheet: {s}")

wb.save(team)
wb2 = openpyxl.load_workbook(team)
print("After:", wb2.sheetnames)
print("Done.")
