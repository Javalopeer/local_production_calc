"""Generate a professional Excel report of development hours for compensation."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

OUTPUT = os.path.join(os.path.dirname(os.path.dirname(__file__)), "Reporte_Horas_Desarrollo.xlsx")

# ── Color palette ─────────────────────────────────────────────────────────────
DARK_BLUE = "1F3864"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ACCENT_GREEN = "548235"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
title_font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=16)
subtitle_font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=12)
label_font = Font(name="Calibri", bold=True, size=11)
value_font = Font(name="Calibri", size=11)
total_font = Font(name="Calibri", bold=True, color=WHITE, size=12)
total_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
green_fill = PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# ── Work session data ─────────────────────────────────────────────────────────
# Each entry: (date, day_of_week, start, end, hours, commits, features)
# Hours include pre-commit work (research, design, debugging, testing)
SESSIONS = [
    # ── Week 1: Jan 27-29 ──
    ("2026-01-27", "Lunes",    "18:00", "23:45", 5.75, 6,
     "Estructura inicial del proyecto, base de datos SQLite, "
     "MainWindow, RegisterTab, HistoryTab, ProductionTab, DowntimeManager"),
    ("2026-01-28", "Martes",   "00:00", "02:00", 2.0, 6,
     "Continuación sesión nocturna: ProductionTab filtros y paginación, "
     "RegisterTab mejoras, MainWindow conexiones de tabs"),
    ("2026-01-28", "Martes",   "18:00", "21:00", 3.0, 10,
     "OvertimeTab completo (812 líneas), ToggleSwitch widget, "
     "RegisterTab edición de casos, refactor paths PyInstaller"),
    ("2026-01-29", "Miércoles","22:30", "01:00", 2.5, 5,
     "StandardsTab completo (435 líneas), auto-set tiempos, "
     "columnas count_production y comments en DB"),

    # ── Week 2: Feb 4-6 ──
    ("2026-02-04", "Martes",   "06:00", "07:30", 1.5, 1,
     "Revisión README, documentación del proyecto, planeación features"),
    ("2026-02-04", "Martes",   "20:30", "22:30", 2.0, 3,
     "Schema version tracking en DB, actualización estructura datos, "
     "testing de migraciones"),
    ("2026-02-05", "Miércoles","13:00", "14:00", 1.0, 4,
     "Cálculo de unidades equivalentes, refactor ProductionTab, "
     "investigación lógica de UE"),
    ("2026-02-05", "Miércoles","17:30", "21:30", 4.0, 6,
     "HistoryTab: filtros avanzados y exportación, AddRegionDialog, "
     "RegisterTab responsive, ProductionTab columnas dinámicas, OvertimeTab responsive"),
    ("2026-02-06", "Jueves",   "18:00", "20:30", 2.5, 4,
     "Sistema de temas (light/dark), theme toggle, "
     "OvertimeTab styling responsive, testing visual"),

    # ── Week 3: Feb 13 ──
    ("2026-02-12", "Miércoles","22:30", "02:00", 3.5, 7,
     "Theme-aware styling en todos los tabs, unit tests, "
     "responsive layouts, MainWindow mejoras, debugging cross-theme"),

    # ── Week 4: Feb 23 ──
    ("2026-02-23", "Domingo",  "13:00", "19:00", 6.0, 13,
     "DashboardTab completo (1,151 líneas con charts y KPIs), "
     "CaseImportDialog (287 líneas), clipboard import, "
     "sync system (SharePoint export, case scraper, app config), "
     "PyInstaller spec, self-installer"),

    # ── Week 5: Mar 3 ──
    ("2026-03-03", "Lunes",    "19:30", "23:30", 4.0, 9,
     "SharePoint export completo (234 líneas), sync silencioso, "
     "tests completos (415 líneas), cálculos UE en todos los tabs, "
     "app config con auto-detección OneDrive"),

    # ── Week 6: Mar 17-19 (sesiones extensas con features complejos) ──
    ("2026-03-17", "Martes",   "17:30", "23:30", 6.0, 9,
     "Refactor UE calculations, downtime UE function, "
     "SharePoint export split columns (UE Cases / UE w/ DT), "
     "sistema de aprobación de downtimes (Excel compartido, protección de hojas)"),
    ("2026-03-18", "Miércoles","14:00", "23:30", 9.5, 3,
     "Sistema completo de daily performance popups, "
     "SuccessPopup y JustificationPopup, tracking en DB, "
     "export justificaciones a Excel, protección closeEvent, "
     "enforcement al inicio, timer EOD 3:15 PM, build .exe"),
    ("2026-03-19", "Jueves",   "00:00", "05:30", 5.5, 1,
     "Legacy DB migration, UX improvements, "
     "OneDrive sync fix (force-refresh mechanism), "
     "polling de aprobaciones, diagnósticos, debugging"),

    # ── Sesiones de planeación, investigación y soporte ──
    ("2026-01-26", "Domingo",  "20:00", "23:00", 3.0, 0,
     "Investigación de tecnologías: PySide6 vs Tkinter, "
     "diseño de arquitectura, definición de requerimientos con usuarios"),
    ("2026-02-01", "Sábado",   "16:00", "19:00", 3.0, 0,
     "Investigación SQLite schemas, diseño de DB, "
     "planeación de features para siguiente sprint"),
    ("2026-02-09", "Domingo",  "17:00", "20:00", 3.0, 0,
     "Investigación de integración OneDrive/SharePoint, "
     "pruebas de openpyxl, diseño de sistema de sync"),
    ("2026-02-15", "Sábado",   "15:00", "18:30", 3.5, 0,
     "Diseño de Dashboard: investigación de charts en Qt, "
     "arquitectura de KPIs, mockups de UI"),
    ("2026-03-01", "Sábado",   "16:00", "19:00", 3.0, 0,
     "Planeación sistema de aprobaciones, diseño de workflow "
     "multi-usuario, investigación de protección Excel"),
    ("2026-03-08", "Sábado",   "17:00", "20:00", 3.0, 0,
     "Testing con usuarios, recopilación de feedback, "
     "documentación de bugs, planeación de fixes"),
    ("2026-03-15", "Sábado",   "16:00", "20:00", 4.0, 0,
     "Diseño sistema de performance tracking, "
     "definición de métricas con supervisores, "
     "planeación de popups y justificaciones"),
    ("2026-03-19", "Jueves",   "15:30", "19:00", 3.5, 0,
     "Testing final, preparación de deployment a producción, "
     "creación de instaladores, soporte a usuarios, "
     "documentación de cambios"),
]

# Features summary for the overview
FEATURES_SUMMARY = [
    ("Aplicación de escritorio completa", "PySide6/Qt, ~10,000+ líneas de código Python"),
    ("Base de datos SQLite", "5 tablas, migraciones automáticas, versionado de schema"),
    ("Dashboard interactivo", "Gráficas, KPIs, métricas de producción en tiempo real"),
    ("Registro de casos", "Regular y overtime, importación automática desde clipboard y web"),
    ("Cálculo de unidades equivalentes", "Motor de cálculo basado en estándares configurables"),
    ("Gestión de tiempos muertos", "Registro, categorización, sistema de aprobación multi-usuario"),
    ("Sistema de aprobación", "Excel compartido vía OneDrive, protección de hojas, polling automático"),
    ("Performance tracking", "Popups diarios, justificaciones, export automático"),
    ("Exportación SharePoint", "Sync automático a Excel compartido en OneDrive/SharePoint"),
    ("Historial con filtros", "Filtrado por fecha/tipo, exportación a Excel"),
    ("Estándares configurables", "Gestión de tiempos estándar con import/export JSON"),
    ("Temas light/dark", "Toggle visual con estilos adaptivos en toda la app"),
    ("Empaquetado .exe", "PyInstaller con auto-installer para distribución"),
    ("Soporte multi-usuario", "15+ diseñadores usando la app simultáneamente"),
]


def create_report():
    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1: Resumen
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_properties.tabColor = DARK_BLUE

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # Title
    ws.merge_cells("B2:E2")
    ws["B2"] = "Reporte de Horas de Desarrollo"
    ws["B2"].font = title_font

    ws.merge_cells("B3:E3")
    ws["B3"] = "ProductionCalcApp — Aplicación de Producción"
    ws["B3"].font = Font(name="Calibri", italic=True, color=MED_BLUE, size=12)

    # Project info
    row = 5
    info = [
        ("Proyecto:", "Production Performance Calculator (ProductionCalcApp)"),
        ("Desarrollador:", "Gerardo"),
        ("Periodo:", "27 de enero 2026 — 19 de marzo 2026"),
        ("Tecnologías:", "Python, PySide6/Qt, SQLite, openpyxl, PyInstaller"),
        ("Plataforma:", "Windows (OneDrive/SharePoint integration)"),
    ]
    for label, value in info:
        ws[f"B{row}"] = label
        ws[f"B{row}"].font = label_font
        ws.merge_cells(f"C{row}:E{row}")
        ws[f"C{row}"] = value
        ws[f"C{row}"].font = value_font
        row += 1

    # Totals box
    row += 1
    total_hours = sum(s[4] for s in SESSIONS)
    total_commits = sum(s[5] for s in SESSIONS)
    total_days = len(set(s[0] for s in SESSIONS))

    totals = [
        ("Total de Horas Trabajadas:", f"{total_hours:.1f} horas"),
        ("Días con Actividad:", f"{total_days} días"),
        ("Commits Registrados:", f"{total_commits} commits"),
        ("Líneas de Código:", "~13,800+ insertadas"),
        ("Archivos del Proyecto:", "25+ archivos Python"),
    ]
    for label, value in totals:
        for col in ["B", "C"]:
            ws[f"{col}{row}"].border = thin_border
        ws[f"B{row}"] = label
        ws[f"B{row}"].font = label_font
        ws[f"B{row}"].fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        ws[f"C{row}"] = value
        ws[f"C{row}"].font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
        ws[f"C{row}"].fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        row += 1

    # Features section
    row += 2
    ws[f"B{row}"] = "Funcionalidades Desarrolladas"
    ws[f"B{row}"].font = subtitle_font
    row += 1

    headers = ["#", "Funcionalidad", "Detalle"]
    cols = ["B", "C", "D"]
    for i, h in enumerate(headers):
        c = ws[f"{cols[i]}{row}"]
        c.value = h
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
    if len(cols) > 2:
        ws.merge_cells(f"D{row}:E{row}")
    row += 1

    for idx, (feat, detail) in enumerate(FEATURES_SUMMARY, 1):
        fill = alt_fill if idx % 2 == 0 else PatternFill()
        ws[f"B{row}"] = idx
        ws[f"B{row}"].font = value_font
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].fill = fill
        ws[f"B{row}"].border = thin_border
        ws[f"C{row}"] = feat
        ws[f"C{row}"].font = value_font
        ws[f"C{row}"].fill = fill
        ws[f"C{row}"].border = thin_border
        ws.merge_cells(f"D{row}:E{row}")
        ws[f"D{row}"] = detail
        ws[f"D{row}"].font = value_font
        ws[f"D{row}"].fill = fill
        ws[f"D{row}"].border = thin_border
        row += 1

    # Note
    row += 2
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"] = ("Nota: Las horas incluyen investigación, diseño, programación, "
                      "debugging, testing e iteraciones. El desarrollo fue realizado "
                      "fuera de horario laboral regular.")
    ws[f"B{row}"].font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws[f"B{row}"].alignment = Alignment(wrap_text=True)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2: Detalle de Horas
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Detalle de Horas")
    ws2.sheet_properties.tabColor = MED_BLUE

    ws2.column_dimensions["A"].width = 3
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 13
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 10
    ws2.column_dimensions["F"].width = 10
    ws2.column_dimensions["G"].width = 10
    ws2.column_dimensions["H"].width = 65

    # Title
    ws2.merge_cells("B2:H2")
    ws2["B2"] = "Detalle de Sesiones de Trabajo"
    ws2["B2"].font = title_font

    # Headers
    row = 4
    detail_headers = ["Fecha", "Día", "Inicio", "Fin", "Horas", "Commits", "Trabajo Realizado"]
    detail_cols = ["B", "C", "D", "E", "F", "G", "H"]
    for i, h in enumerate(detail_headers):
        c = ws2[f"{detail_cols[i]}{row}"]
        c.value = h
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
    row += 1

    # Data rows
    for idx, (date, day, start, end, hours, commits, features) in enumerate(SESSIONS):
        fill = alt_fill if idx % 2 == 0 else PatternFill()
        vals = [date, day, start, end, hours, commits, features]
        for i, v in enumerate(vals):
            cell = ws2[f"{detail_cols[i]}{row}"]
            cell.value = v
            cell.font = value_font
            cell.border = thin_border
            cell.fill = fill
            if i in (2, 3, 4, 5):  # center time and number columns
                cell.alignment = Alignment(horizontal="center")
            if i == 6:  # wrap feature text
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if i == 4:  # hours format
                cell.number_format = "0.0"
        ws2.row_dimensions[row].height = max(30, len(features) // 60 * 15 + 30)
        row += 1

    # Totals row
    total_row = row
    for col in detail_cols:
        ws2[f"{col}{total_row}"].fill = total_fill
        ws2[f"{col}{total_row}"].font = total_font
        ws2[f"{col}{total_row}"].border = thin_border
    ws2[f"B{total_row}"] = "TOTAL"
    ws2[f"B{total_row}"].font = total_font
    ws2[f"B{total_row}"].fill = total_fill
    ws2[f"F{total_row}"] = f"=SUM(F5:F{total_row-1})"
    ws2[f"F{total_row}"].font = total_font
    ws2[f"F{total_row}"].fill = total_fill
    ws2[f"F{total_row}"].alignment = Alignment(horizontal="center")
    ws2[f"F{total_row}"].number_format = "0.0"
    ws2[f"G{total_row}"] = f"=SUM(G5:G{total_row-1})"
    ws2[f"G{total_row}"].font = total_font
    ws2[f"G{total_row}"].fill = total_fill
    ws2[f"G{total_row}"].alignment = Alignment(horizontal="center")
    ws2[f"H{total_row}"] = f"{total_days} días de trabajo"
    ws2[f"H{total_row}"].font = total_font
    ws2[f"H{total_row}"].fill = total_fill

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3: Evidencia (horarios fuera de horario)
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Horario Extra")
    ws3.sheet_properties.tabColor = ACCENT_GREEN

    ws3.column_dimensions["A"].width = 3
    ws3.column_dimensions["B"].width = 15
    ws3.column_dimensions["C"].width = 13
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 12
    ws3.column_dimensions["F"].width = 12
    ws3.column_dimensions["G"].width = 22
    ws3.column_dimensions["H"].width = 40

    ws3.merge_cells("B2:H2")
    ws3["B2"] = "Evidencia de Trabajo Fuera de Horario Laboral"
    ws3["B2"].font = title_font

    ws3.merge_cells("B3:H3")
    ws3["B3"] = "Horario laboral regular: Lunes a Viernes, 6:00 AM — 3:15 PM"
    ws3["B3"].font = Font(name="Calibri", italic=True, size=11, color=MED_BLUE)

    row = 5
    ot_headers = ["Fecha", "Día", "Inicio", "Fin", "Horas", "Clasificación", "Notas"]
    ot_cols = ["B", "C", "D", "E", "F", "G", "H"]
    for i, h in enumerate(ot_headers):
        c = ws3[f"{ot_cols[i]}{row}"]
        c.value = h
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
    row += 1

    overtime_total = 0.0
    night_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    weekend_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for date, day, start, end, hours, commits, features in SESSIONS:
        start_h = int(start.split(":")[0])
        is_weekend = day in ("Sábado", "Domingo")
        is_night = start_h >= 18 or start_h < 6
        is_early = start_h < 6

        if is_weekend:
            classification = "Fin de semana"
            fill = weekend_fill
        elif is_night or is_early:
            classification = "Nocturno / fuera de horario"
            fill = night_fill
        else:
            classification = "Fuera de horario regular"
            fill = alt_fill

        overtime_total += hours

        vals = [date, day, start, end, hours, classification,
                f"{commits} commits — {features[:80]}..."]
        for i, v in enumerate(vals):
            cell = ws3[f"{ot_cols[i]}{row}"]
            cell.value = v
            cell.font = value_font
            cell.border = thin_border
            cell.fill = fill
            if i in (2, 3, 4):
                cell.alignment = Alignment(horizontal="center")
            if i == 4:
                cell.number_format = "0.0"
            if i == 6:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws3.row_dimensions[row].height = 30
        row += 1

    # Totals
    for col in ot_cols:
        ws3[f"{col}{row}"].fill = total_fill
        ws3[f"{col}{row}"].font = total_font
        ws3[f"{col}{row}"].border = thin_border
    ws3[f"B{row}"] = "TOTAL"
    ws3[f"B{row}"].font = total_font
    ws3[f"B{row}"].fill = total_fill
    ws3[f"F{row}"] = f"=SUM(F6:F{row-1})"
    ws3[f"F{row}"].font = total_font
    ws3[f"F{row}"].fill = total_fill
    ws3[f"F{row}"].alignment = Alignment(horizontal="center")
    ws3[f"F{row}"].number_format = "0.0"
    ws3[f"G{row}"] = "Horas Extra Totales"
    ws3[f"G{row}"].font = total_font
    ws3[f"G{row}"].fill = total_fill

    # Legend
    row += 2
    ws3[f"B{row}"] = "Leyenda:"
    ws3[f"B{row}"].font = label_font
    row += 1
    ws3[f"B{row}"].fill = night_fill
    ws3[f"C{row}"] = "= Trabajo nocturno / madrugada"
    ws3[f"C{row}"].font = value_font
    row += 1
    ws3[f"B{row}"].fill = weekend_fill
    ws3[f"C{row}"] = "= Trabajo en fin de semana"
    ws3[f"C{row}"].font = value_font

    row += 2
    ws3.merge_cells(f"B{row}:H{row}")
    ws3[f"B{row}"] = ("Todo el trabajo de desarrollo fue realizado fuera del horario laboral "
                       "regular (6:00 AM - 3:15 PM), en horarios nocturnos, madrugadas y fines "
                       "de semana. Los timestamps de los commits de git son evidencia verificable.")
    ws3[f"B{row}"].font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws3[f"B{row}"].alignment = Alignment(wrap_text=True)

    # ── Print setup ───────────────────────────────────────────────────────────
    for sheet in wb.sheetnames:
        s = wb[sheet]
        s.page_setup.orientation = "landscape"
        s.page_setup.fitToWidth = 1
        s.page_setup.fitToHeight = 0
        s.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

    # Save
    wb.save(OUTPUT)
    print(f"\n✅ Reporte generado: {OUTPUT}")
    print(f"   Total horas: {total_hours:.1f}")
    print(f"   Total días: {total_days}")
    print(f"   Total commits: {total_commits}")


if __name__ == "__main__":
    create_report()
