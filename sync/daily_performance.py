# -*- coding: utf-8 -*-
"""
Daily Performance Check — end-of-day popups and justification export.

Logic:
  1. After each case save, check if production ≥ 95% OR UE ≥ 14.
     → If yes, show a congratulations popup with daily summary.
  2. At 3:05 PM, if production < 95% AND UE < 14:
     → Show an encouragement popup that requires a written justification.
     → The app cannot be closed until the justification is submitted.
  3. If the designer exits before 3:05 PM *without* meeting the target,
     the next day the app blocks usage until the justification is submitted.
  4. Justifications are exported to a shared Excel for supervisor review.
"""

import os
import time
from datetime import datetime, date

from db.database import get_connection
from sync.app_config import load_config
from tabs.utils import (
    DAILY_BASE_MINUTES,
    load_units_eq_data,
    calculate_equivalent_units,
    calculate_downtime_equivalent_units,
)

PRODUCTION_TARGET_PCT = 95.0
UE_TARGET = 14.0

SHARED_JUSTIFICATION_FILE = "_Daily_Justifications.xlsx"

_OPENPYXL_OK = False
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
    _OPENPYXL_OK = True
except ImportError:
    pass


# ── helpers ──────────────────────────────────────────────────────────────────

def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _get_justification_path() -> str | None:
    """Return the full path to the shared justification Excel."""
    cfg = load_config()
    folder = cfg.get("export_folder", "").strip()
    if not folder or not os.path.isdir(folder):
        return None
    dt_dir = os.path.join(folder, "Downtime")
    try:
        os.makedirs(dt_dir, exist_ok=True)
    except OSError:
        return None
    return os.path.join(dt_dir, SHARED_JUSTIFICATION_FILE)


# ── DB table for tracking ────────────────────────────────────────────────────

def init_performance_table():
    """Create the daily_performance table if it doesn't exist."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS daily_performance (
            fecha TEXT PRIMARY KEY,
            production_pct REAL DEFAULT 0,
            equivalent_units REAL DEFAULT 0,
            met_target INTEGER DEFAULT 0,
            justification TEXT DEFAULT '',
            justification_submitted INTEGER DEFAULT 0,
            popup_shown INTEGER DEFAULT 0
        )
    """)
    conn.commit()
    conn.close()


def _ensure_row(fecha: str):
    """Ensure a row exists for the given date."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT OR IGNORE INTO daily_performance (fecha) VALUES (?)",
        (fecha,),
    )
    conn.commit()
    conn.close()


# ── Daily metrics query ──────────────────────────────────────────────────────

def get_daily_metrics(fecha: str) -> dict:
    """Compute production %, UE, cases count, and downtime for a given date.

    Returns a dict with keys:
      production_pct, cases_pct, downtime_pct, equivalent_units,
      cases_ue, downtime_ue, total_cases, total_downtime_min,
      met_target, case_breakdown (list of (region, type, count, cv))
    """
    conn = get_connection()
    cur = conn.cursor()

    # Regular cases (count_production = 1)
    cur.execute("""
        SELECT SUM(case_value) FROM cases
        WHERE fecha = ? AND (count_production = 1 OR count_production IS NULL)
    """, (fecha,))
    cases_pct = cur.fetchone()[0] or 0.0

    # Count of cases
    cur.execute("""
        SELECT COUNT(*) FROM cases
        WHERE fecha = ? AND (count_production = 1 OR count_production IS NULL)
    """, (fecha,))
    total_cases = cur.fetchone()[0] or 0

    # Cases by region+type for UE
    cur.execute("""
        SELECT region, tipo_caso, COUNT(*), SUM(case_value)
        FROM cases
        WHERE fecha = ? AND (count_production = 1 OR count_production IS NULL)
        GROUP BY region, tipo_caso
    """, (fecha,))
    case_breakdown = cur.fetchall()

    # Downtime
    cur.execute("""
        SELECT SUM(duracion) FROM downtimes
        WHERE fecha = ? AND (status = 'approved' OR status IS NULL)
    """, (fecha,))
    total_downtime_min = cur.fetchone()[0] or 0.0

    conn.close()

    # Calculate UE
    units_eq = load_units_eq_data()
    cases_ue = 0.0
    for region, case_type, count, sum_cv in case_breakdown:
        if count and sum_cv:
            cases_ue += calculate_equivalent_units(
                units_eq, region or "", case_type or "", sum_cv or 0.0, count=count or 0
            )

    downtime_pct = (total_downtime_min / DAILY_BASE_MINUTES) * 100 if total_downtime_min > 0 else 0.0
    downtime_ue = calculate_downtime_equivalent_units(total_downtime_min)

    production_pct = cases_pct + downtime_pct
    equivalent_units = cases_ue + downtime_ue
    met_target = production_pct >= PRODUCTION_TARGET_PCT or equivalent_units >= UE_TARGET

    return {
        "production_pct": production_pct,
        "cases_pct": cases_pct,
        "downtime_pct": downtime_pct,
        "equivalent_units": equivalent_units,
        "cases_ue": cases_ue,
        "downtime_ue": downtime_ue,
        "total_cases": total_cases,
        "total_downtime_min": total_downtime_min,
        "met_target": met_target,
        "case_breakdown": case_breakdown,
    }


# ── State queries ────────────────────────────────────────────────────────────

def has_pending_justification() -> str | None:
    """Check if there is an older date where the target was not met and no
    justification was submitted.  Returns the date string, or None.
    Skips weekends and days with zero production (person wasn't producing)."""
    today = date.today().isoformat()
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT fecha, production_pct, equivalent_units FROM daily_performance
        WHERE fecha < ? AND met_target = 0 AND justification_submitted = 0
        ORDER BY fecha DESC LIMIT 1
    """, (today,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return None
    fecha_str, prod_pct, eq_units = row[0], row[1] or 0, row[2] or 0
    # Skip if the pending date was a weekend
    try:
        pending = date.fromisoformat(fecha_str)
        if pending.weekday() >= 5:
            return None
    except (ValueError, TypeError):
        pass
    # Skip if zero production (person wasn't producing that day)
    if prod_pct == 0 and eq_units == 0:
        return None
    return fecha_str


def record_daily_snapshot(fecha: str, metrics: dict):
    """Save daily metrics snapshot to the DB."""
    _ensure_row(fecha)
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE daily_performance
        SET production_pct = ?, equivalent_units = ?, met_target = ?
        WHERE fecha = ?
    """, (metrics["production_pct"], metrics["equivalent_units"],
          1 if metrics["met_target"] else 0, fecha))
    conn.commit()
    conn.close()


def mark_success_shown(fecha: str):
    """Record that the success popup was shown for this date."""
    _ensure_row(fecha)
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "UPDATE daily_performance SET popup_shown = 1 WHERE fecha = ?",
        (fecha,),
    )
    conn.commit()
    conn.close()


def was_success_shown(fecha: str) -> bool:
    """Check if the success popup was already shown for this date."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT popup_shown FROM daily_performance WHERE fecha = ?",
        (fecha,),
    )
    row = cur.fetchone()
    conn.close()
    return bool(row and row[0])


def save_justification(fecha: str, text: str):
    """Save the justification text and mark as submitted."""
    _ensure_row(fecha)
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE daily_performance
        SET justification = ?, justification_submitted = 1
        WHERE fecha = ?
    """, (text, fecha))
    conn.commit()
    conn.close()


# ── Excel export ─────────────────────────────────────────────────────────────

def _get_over_standard_cases(fecha: str) -> list:
    """Return cases where tiempo_real > std_time for the given date.

    Each row: (case_id, region, tipo_caso, doctor, hora_inicio, hora_fin,
               tiempo_real, std_time, difference)
    """
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT case_id, region, tipo_caso, doctor,
               hora_inicio, hora_fin, tiempo_real, std_time
        FROM cases
        WHERE fecha = ?
          AND (count_production = 1 OR count_production IS NULL)
          AND tiempo_real > std_time
          AND std_time > 0
        ORDER BY (tiempo_real - std_time) DESC
    """, (fecha,))
    rows = cur.fetchall()
    conn.close()

    result = []
    for r in rows:
        case_id, region, tipo, doctor, h_ini, h_fin, t_real, t_std = r
        diff = round((t_real or 0) - (t_std or 0), 2)
        result.append((case_id, region, tipo, doctor or "",
                        h_ini or "", h_fin or "",
                        round(t_real or 0, 2), round(t_std or 0, 2), diff))
    return result


def export_justification(designer_name: str, fecha: str, metrics: dict, justification: str) -> bool:
    """Append justification to the shared Excel file."""
    if not _OPENPYXL_OK:
        return False

    path = _get_justification_path()
    if not path:
        return False

    _locked = Protection(locked=True)

    headers = [
        "Designer", "Date", "Production (%)", "UE", "Cases",
        "Downtime (min)", "Justification", "Submitted At",
    ]
    hdr_fill = PatternFill("solid", fgColor="E65100")
    hdr_font = Font(color="FFFFFF", bold=True)
    row_fill = PatternFill("solid", fgColor="FFF3E0")

    # Load existing or create new
    existing_rows = []
    existing_over = []
    if os.path.exists(path):
        try:
            wb_old = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws_old = wb_old.active
            for r in ws_old.iter_rows(min_row=2, values_only=True):
                if r and r[0] is not None:
                    existing_rows.append(r)
            # Read existing over-standard rows (other designers/dates)
            if "Over-Standard Cases" in wb_old.sheetnames:
                ws_old2 = wb_old["Over-Standard Cases"]
                for r in ws_old2.iter_rows(min_row=2, values_only=True):
                    if r and r[0] is not None:
                        r_designer = str(r[0] or "").strip().lower()
                        r_date = str(r[1] or "").strip()
                        if r_designer != designer_name.strip().lower() or r_date != fecha:
                            existing_over.append(r)
            wb_old.close()
        except Exception:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Justifications"
    ws.sheet_view.showGridLines = False

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin()
        cell.protection = _locked

    # Write existing rows
    row_idx = 2
    for r in existing_rows:
        for c_idx, val in enumerate(r, 1):
            cell = ws.cell(row_idx, c_idx, val)
            cell.alignment = Alignment(
                horizontal="left" if c_idx in (1, 7) else "center",
                vertical="center",
            )
            cell.fill = row_fill
            cell.border = _thin()
            cell.protection = _locked
        row_idx += 1

    # Append new justification
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    new_row = [
        designer_name,
        fecha,
        round(metrics.get("production_pct", 0), 2),
        round(metrics.get("equivalent_units", 0), 2),
        metrics.get("total_cases", 0),
        round(metrics.get("total_downtime_min", 0), 1),
        justification,
        now_str,
    ]
    for c_idx, val in enumerate(new_row, 1):
        cell = ws.cell(row_idx, c_idx, val)
        cell.alignment = Alignment(
            horizontal="left" if c_idx in (1, 7) else "center",
            vertical="center",
        )
        cell.fill = row_fill
        cell.border = _thin()
        cell.protection = _locked

    # Column widths
    widths = {"A": 22, "B": 14, "C": 16, "D": 10, "E": 10, "F": 16, "G": 60, "H": 18}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{row_idx}"

    # Protect entire sheet — read-only but allow filtering
    ws.protection.sheet = True
    ws.protection.password = "spark2026"
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.enable()

    # ── Sheet 2: Over-standard cases ─────────────────────────────────────────
    over_cases = _get_over_standard_cases(fecha)

    ws2 = wb.create_sheet("Over-Standard Cases")
    ws2.sheet_view.showGridLines = False

    over_headers = [
        "Designer", "Date", "Case ID", "Region", "Case Type",
        "Doctor", "Start", "End", "Actual (min)", "Standard (min)", "Over (min)",
    ]
    over_hdr_fill = PatternFill("solid", fgColor="C62828")
    for col, h in enumerate(over_headers, 1):
        cell = ws2.cell(1, col, h)
        cell.fill = over_hdr_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin()
        cell.protection = _locked

    ov_row_fill = PatternFill("solid", fgColor="FFEBEE")
    ov_idx = 2

    # Write existing over-standard rows (other designers/dates)
    for r in existing_over:
        for c, val in enumerate(r, 1):
            cell = ws2.cell(ov_idx, c, val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = ov_row_fill
            cell.border = _thin()
            cell.protection = _locked
        ov_idx += 1

    # Write current designer's over-standard cases
    for case_row in over_cases:
        case_id, region, tipo, doctor, h_ini, h_fin, t_real, t_std, diff = case_row
        vals = [designer_name, fecha, case_id, region, tipo,
                doctor, h_ini, h_fin, t_real, t_std, diff]
        for c, val in enumerate(vals, 1):
            cell = ws2.cell(ov_idx, c, val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = ov_row_fill
            cell.border = _thin()
            cell.protection = _locked
        ov_idx += 1

    # Column widths for over-standard sheet
    ov_widths = {"A": 22, "B": 14, "C": 14, "D": 16, "E": 16,
                 "F": 22, "G": 10, "H": 10, "I": 14, "J": 14, "K": 14}
    for col_letter, width in ov_widths.items():
        ws2.column_dimensions[col_letter].width = width
    ws2.row_dimensions[1].height = 22
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:K{ov_idx}"

    ws2.protection.sheet = True
    ws2.protection.password = "spark2026"
    ws2.protection.autoFilter = False
    ws2.protection.sort = False
    ws2.protection.enable()

    # Save with retry
    for attempt in range(3):
        try:
            wb.save(path)
            return True
        except PermissionError:
            if attempt < 2:
                time.sleep(1)
        except Exception:
            return False
    return False
