# -*- coding: utf-8 -*-
"""
SharePoint / OneDrive Excel export.

Generates an .xlsx file with:
  Sheet 1 – Daily Summary    (today's production overview)
  Sheet 2 – Cases Today      (all cases recorded today)
  Sheet 3 – Monthly Summary  (daily totals for the current month)

The file is saved to the configured export_folder so OneDrive
automatically syncs it to SharePoint.

Filename pattern:
  <DesignerName>_Production_<YYYY-MM-DD>.xlsx
"""
import os
import random
import sqlite3
import time
from datetime import datetime, date

_OPENPYXL_ERROR = ""
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.filters import FilterColumn, Filters
    _OPENPYXL_OK = True
except Exception as _e:
    _OPENPYXL_OK = False
    _OPENPYXL_ERROR = str(_e)

from db.database import DB_PATH
from sync.app_config import load_config
from sync.app_logger import log_event
from tabs.utils import DAILY_BASE_MINUTES

# ── Colour palette ──────────────────────────────────────────────────────────
_BLUE       = "2D89EF"
_BLUE_LIGHT = "D6E8FF"
_GREEN      = "4CAF50"
_YELLOW     = "FFC107"
_RED        = "F44336"
_HEADER_FG  = "FFFFFF"
_GREY_ROW   = "F5F5F5"
_TITLE_GREY = "3C3C3C"

# Cache: track latest mtime of _Summary_*.xlsx to skip unnecessary Dashboard rebuilds
_last_dashboard_max_mtime: float = 0.0


# ── Helpers ─────────────────────────────────────────────────────────────────

def _cleanup_onedrive_conflicts(productions_dir: str):
    """Remove OneDrive conflict copies like _Dashboard-CRI-MACHINE.xlsx.

    OneDrive creates these when two machines write the same file before sync
    completes.  The main _Dashboard.xlsx is always the authoritative version.
    """
    import re
    try:
        for f in os.listdir(productions_dir):
            # Match patterns like _Dashboard-CRI-LGONZALEZA.xlsx or _Summary_Name-MACHINE.xlsx
            if re.match(r"_Dashboard-.+\.xlsx$", f) or re.match(r".*-[A-Z]{2,5}-[A-Z]+\.xlsx$", f):
                fpath = os.path.join(productions_dir, f)
                try:
                    os.remove(fpath)
                    print(f"[sharepoint_sync] Removed OneDrive conflict copy: {f}")
                except OSError as exc:
                    log_event("sharepoint_sync", f"failed removing conflict copy {f}: {exc}", level="WARN")
        # Also check Dashboards/ subfolder
        dashboards_dir = os.path.join(productions_dir, "Dashboards")
        if os.path.isdir(dashboards_dir):
            for f in os.listdir(dashboards_dir):
                if re.match(r"_Dashboard_.+-[A-Z]{2,5}-[A-Z]+\.xlsx$", f):
                    try:
                        os.remove(os.path.join(dashboards_dir, f))
                        print(f"[sharepoint_sync] Removed conflict copy: Dashboards/{f}")
                    except OSError as exc:
                        log_event("sharepoint_sync", f"failed removing dashboard conflict copy {f}: {exc}", level="WARN")
    except Exception as exc:
        log_event("sharepoint_sync", f"conflict cleanup failed: {exc}", level="WARN")


def _db():
    return sqlite3.connect(DB_PATH)

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_fill(hex_color=_BLUE):
    return PatternFill("solid", fgColor=hex_color)

def _row_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _hdr(cell, text, bold=True, color=_HEADER_FG, bg=_BLUE, size=11, wrap=False):
    cell.value = text
    cell.font = Font(bold=bold, color=color, size=size)
    cell.fill = _header_fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = _thin_border()

def _cell(cell, value, bold=False, color="000000", align="left", bg=None, num_fmt=None):
    cell.value = value
    cell.font = Font(bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _thin_border()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        cell.number_format = num_fmt

def _autowidth(ws, extra=4):
    max_col = ws.max_column or 1
    for ci in range(1, max_col + 1):
        length = 0
        for row in ws.iter_rows(min_col=ci, max_col=ci):
            for cell in row:
                try:
                    if cell.value:
                        length = max(length, len(str(cell.value)))
                except Exception as exc:
                    log_event("sharepoint_sync", f"autowidth value read failed col={ci}: {exc}", level="WARN")
        ws.column_dimensions[get_column_letter(ci)].width = max(length, 8) + extra

def _production_color(pct):
    if pct >= 100:  return _GREEN
    if pct >= 95:   return _YELLOW
    return _RED


# ── Data queries ─────────────────────────────────────────────────────────────

def _get_daily_data(target_date: str):
    conn = _db()
    cur  = conn.cursor()

    # Cases
    cur.execute("""
        SELECT case_id, region, tipo_caso, doctor, hora_inicio, hora_fin,
               std_time, tiempo_real, efficiency, estado, case_value, count_production, comments
        FROM cases WHERE fecha = ? ORDER BY hora_inicio
    """, (target_date,))
    cases = cur.fetchall()

    # OT Cases
    cur.execute("""
        SELECT case_id, region, tipo_caso, doctor, hora_inicio, hora_fin,
               std_time, tiempo_real, efficiency, estado, case_value, count_production, comments
        FROM ot_cases WHERE fecha = ? ORDER BY hora_inicio
    """, (target_date,))
    ot_cases = cur.fetchall()

    # Downtime
    cur.execute("""
        SELECT hora_inicio, hora_fin, duracion, razon
        FROM downtimes WHERE fecha = ? ORDER BY hora_inicio
    """, (target_date,))
    downtimes = cur.fetchall()

    # Totals (only count_production = 1)
    cur.execute("""
        SELECT SUM(case_value) FROM cases
        WHERE fecha = ? AND (count_production = 1 OR count_production IS NULL)
    """, (target_date,))
    total_cases_pct = cur.fetchone()[0] or 0.0

    cur.execute("SELECT SUM(duracion) FROM downtimes WHERE fecha = ? AND (status = 'approved' OR status IS NULL)", (target_date,))
    total_downtime_min = cur.fetchone()[0] or 0.0

    conn.close()
    return cases, ot_cases, downtimes, total_cases_pct, total_downtime_min


def _get_monthly_data(year: int, month: int):
    conn = _db()
    cur  = conn.cursor()
    prefix = f"{year:04d}-{month:02d}-%"

    cur.execute("""
        SELECT fecha, SUM(case_value), COUNT(*)
        FROM cases
        WHERE fecha LIKE ? AND (count_production = 1 OR count_production IS NULL)
        GROUP BY fecha ORDER BY fecha
    """, (prefix,))
    monthly_cases = {r[0]: (r[1] or 0.0, r[2]) for r in cur.fetchall()}

    cur.execute("""
        SELECT fecha, SUM(duracion)
        FROM downtimes WHERE fecha LIKE ? AND (status = 'approved' OR status IS NULL)
        GROUP BY fecha ORDER BY fecha
    """, (prefix,))
    monthly_dt = {r[0]: r[1] or 0.0 for r in cur.fetchall()}

    conn.close()

    # Merge all dates
    all_dates = sorted(set(list(monthly_cases.keys()) + list(monthly_dt.keys())))
    rows = []
    for d in all_dates:
        cases_pct, n_cases = monthly_cases.get(d, (0.0, 0))
        dt_min   = monthly_dt.get(d, 0.0)
        dt_pct   = (dt_min / DAILY_BASE_MINUTES) * 100
        total    = cases_pct + dt_pct
        rows.append((d, n_cases, cases_pct, dt_min, dt_pct, total))
    return rows


_SP_UPLOAD_JITTER = (1.0, 6.0)  # random backoff range (s) between SharePoint write retries

# Fixed case types shown as individual columns in the Dashboard
FIXED_CASE_TYPES = ["Primary", "Secondary", "CR"]

# Full explicit type list for dashboard/history exports (REG + OT)
CASE_TYPE_COLUMNS = [
    "Primary",
    "Secondary",
    "CR",
    "Stage RX Primary",
    "Stage RX Secondary",
    "Stage RX CR",
    "Bite Sync Primary",
    "Bite Sync Secondary",
    "Bite Sync CR",
    "New Impressions",
]


def _bucket_case_types(type_counts: dict) -> dict:
    """Expand raw case-type counts into explicit dashboard columns.

    Returns (counts_by_case_type, other_count) where counts_by_case_type has
    one entry for each value in CASE_TYPE_COLUMNS.
    """
    counts = {name: 0 for name in CASE_TYPE_COLUMNS}
    other_count = 0

    alias_map = {
        "bitesync2": "Bite Sync Secondary",
        "bite sync2": "Bite Sync Secondary",
        "bitesync3": "Bite Sync CR",
        "bite sync3": "Bite Sync CR",
        "new impressions": "New Impressions",
        "newimpressions": "New Impressions",
        "new_impressions": "New Impressions",
        "new impression": "New Impressions",
    }

    for raw_type, count in (type_counts or {}).items():
        n = int(count or 0)
        t_raw = (raw_type or "").strip()
        t = t_raw.lower()

        if not t:
            other_count += n
            continue

        canonical = None
        if t_raw in counts:
            canonical = t_raw
        elif t in alias_map:
            canonical = alias_map[t]
        elif "stage rx" in t:
            if "primary" in t:
                canonical = "Stage RX Primary"
            elif "secondary" in t:
                canonical = "Stage RX Secondary"
            elif t.endswith("cr") or " cr" in t:
                canonical = "Stage RX CR"
        elif "bite sync" in t:
            if "primary" in t:
                canonical = "Bite Sync Primary"
            elif "secondary" in t:
                canonical = "Bite Sync Secondary"
            elif t.endswith("cr") or " cr" in t:
                canonical = "Bite Sync CR"
        elif t == "primary":
            canonical = "Primary"
        elif t == "secondary":
            canonical = "Secondary"
        elif t == "cr":
            canonical = "CR"

        if canonical and canonical in counts:
            counts[canonical] += n
        else:
            other_count += n

    return counts, other_count


def _get_ue_and_types(target_date: str):
    """
    Returns (ue_total, reg_type_counts, ot_type_counts) for target_date.
    reg_type_counts / ot_type_counts are dicts  {tipo_caso: count}.
        UE supports both models:
            - per-type UE (units_eq[region][tipo])
            - legacy base-rate UE (units_eq[region]['100'])
    """
    import json as _json
    try:
        _ueq_path = os.path.join(os.path.dirname(__file__),
                                  "..", "data", "units_eq.json")
        with open(_ueq_path, encoding="utf-8-sig") as _f:
            units_eq = _json.load(_f)
    except Exception as exc:
        log_event("sharepoint_sync", f"units_eq load failed; using empty map: {exc}", level="WARN")
        units_eq = {}

    conn = _db()
    cur  = conn.cursor()

    # UE and reg type counts
    cur.execute("""
        SELECT region, tipo_caso, case_value, count_production
        FROM cases WHERE fecha = ?
    """, (target_date,))
    reg_rows = cur.fetchall()

    # OT type counts
    cur.execute("""
        SELECT tipo_caso FROM ot_cases WHERE fecha = ?
    """, (target_date,))
    ot_rows = cur.fetchall()

    # Downtime total for UE conversion (only approved downtimes count)
    cur.execute("SELECT SUM(duracion) FROM downtimes WHERE fecha = ? AND (status = 'approved' OR status IS NULL)", (target_date,))
    downtime_total_min = cur.fetchone()[0] or 0.0
    conn.close()

    ue_cases = 0.0
    reg_type_counts: dict = {}
    for region, tipo, case_value, count_prod in reg_rows:
        if count_prod in (1, None):
            reg_map = units_eq.get(region, {}) if isinstance(units_eq.get(region, {}), dict) else {}
            if tipo in reg_map:
                try:
                    ue_cases += float(reg_map.get(tipo) or 0.0)
                except (TypeError, ValueError):
                    pass
            else:
                try:
                    rate = float(reg_map.get("100") or 0.0)
                except (TypeError, ValueError):
                    rate = 0.0
                ue_cases += (case_value or 0) * rate / 100.0
        reg_type_counts[tipo] = reg_type_counts.get(tipo, 0) + 1

    ot_type_counts: dict = {}
    for (tipo,) in ot_rows:
        ot_type_counts[tipo] = ot_type_counts.get(tipo, 0) + 1

    ue_total = ue_cases  # downtime no longer contributes to equivalent units

    return ue_cases, ue_total, reg_type_counts, ot_type_counts


# ── Sheet builders ───────────────────────────────────────────────────────────

def _build_daily_summary(wb, target_date: str, designer: str,
                         cases, ot_cases, downtimes,
                         total_cases_pct, total_downtime_min,
                         ue_cases: float = 0.0, ue_total: float = 0.0):
    ws = wb.active
    ws.title = "Daily Summary"
    ws.sheet_view.showGridLines = False

    dt_pct   = (total_downtime_min / DAILY_BASE_MINUTES) * 100
    total_pct = total_cases_pct + dt_pct
    pct_color = _production_color(total_pct)

    # ── Title block ─────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "Production Performance Report"
    t.font  = Font(bold=True, size=15, color=_HEADER_FG)
    t.fill  = _header_fill(_TITLE_GREY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = f"Designer: {designer}     Date: {target_date}"
    s.font  = Font(size=11, color="555555")
    s.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20
    ws["A3"].value = ""

    # ── KPI row ─────────────────────────────────────────────────────
    row = 4
    ws.row_dimensions[row].height = 22
    for col, label in enumerate(["Cases Production", "Downtime", "Total Production",
                                  "Reg Cases", "OT Cases", "Downtime Events",
                                  "UE (Cases)", "UE (w/ Downtime)"], 1):
        _hdr(ws.cell(row, col), label, bg=_BLUE)

    row = 5
    ws.row_dimensions[row].height = 22
    _cell(ws.cell(row, 1), f"{total_cases_pct:.2f}%", bold=True, align="center", bg="DCEEFB")
    _cell(ws.cell(row, 2), f"{total_downtime_min:.0f} min  ({dt_pct:.2f}%)", align="center", bg="DCEEFB")
    _cell(ws.cell(row, 3), f"{total_pct:.2f}%", bold=True, align="center",
          bg=pct_color, color=_HEADER_FG)
    _cell(ws.cell(row, 4), len(cases),    align="center", bg="DCEEFB")
    _cell(ws.cell(row, 5), len(ot_cases), align="center", bg="DCEEFB")
    _cell(ws.cell(row, 6), len(downtimes),align="center", bg="DCEEFB")
    _cell(ws.cell(row, 7), f"{ue_cases:.2f}", bold=True, align="center", bg="DCEEFB")
    _cell(ws.cell(row, 8), f"{ue_total:.2f}", bold=True, align="center", bg="DCEEFB")

    ws["A6"].value = ""

    # ── Cases table ─────────────────────────────────────────────────
    row = 7
    ws.merge_cells(f"A{row}:H{row}")
    t2 = ws.cell(row, 1)
    t2.value = "Regular Cases"
    t2.font  = Font(bold=True, size=11, color=_HEADER_FG)
    t2.fill  = _header_fill("1565C0")
    t2.alignment = Alignment(horizontal="left")

    row = 8
    cols_c = ["Case ID", "Region", "Type", "Start", "End", "Value (%)"]
    for ci, label in enumerate(cols_c, 1):
        _hdr(ws.cell(row, ci), label, bg=_BLUE_LIGHT, color="000000")

    for i, c in enumerate(cases):
        row += 1
        bg = _GREY_ROW if i % 2 == 0 else "FFFFFF"
        _cell(ws.cell(row, 1), c[0],  bg=bg)
        _cell(ws.cell(row, 2), c[1],  bg=bg)
        _cell(ws.cell(row, 3), c[2],  bg=bg)
        _cell(ws.cell(row, 4), c[4],  bg=bg, align="center")
        _cell(ws.cell(row, 5), c[5],  bg=bg, align="center")
        _cell(ws.cell(row, 6), f"{c[10]:.3f}%" if c[10] else "—",
              bg=bg, align="right")

    if not cases:
        row += 1
        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row, 1).value = "No regular cases for this date."

    row += 1; ws.cell(row, 1).value = ""

    # ── OT cases table ──────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    t3 = ws.cell(row, 1)
    t3.value = "Overtime Cases"
    t3.font  = Font(bold=True, size=11, color=_HEADER_FG)
    t3.fill  = _header_fill("6A1B9A")
    t3.alignment = Alignment(horizontal="left")

    row += 1
    for ci, label in enumerate(cols_c, 1):
        _hdr(ws.cell(row, ci), label, bg="F3E5F5", color="000000")

    for i, c in enumerate(ot_cases):
        row += 1
        bg = _GREY_ROW if i % 2 == 0 else "FFFFFF"
        _cell(ws.cell(row, 1), c[0], bg=bg)
        _cell(ws.cell(row, 2), c[1], bg=bg)
        _cell(ws.cell(row, 3), c[2], bg=bg)
        _cell(ws.cell(row, 4), c[4], bg=bg, align="center")
        _cell(ws.cell(row, 5), c[5], bg=bg, align="center")
        _cell(ws.cell(row, 6), f"{c[10]:.3f}%" if c[10] else "—",
              bg=bg, align="right")

    if not ot_cases:
        row += 1
        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row, 1).value = "No OT cases for this date."

    row += 1; ws.cell(row, 1).value = ""

    # ── Downtime table ──────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    t4 = ws.cell(row, 1)
    t4.value = "Downtime"
    t4.font  = Font(bold=True, size=11, color=_HEADER_FG)
    t4.fill  = _header_fill("E65100")
    t4.alignment = Alignment(horizontal="left")

    row += 1
    for ci, label in enumerate(["Start", "End", "Duration (min)", "Reason", "", ""], 1):
        _hdr(ws.cell(row, ci), label, bg="FBE9E7", color="000000")

    for i, d in enumerate(downtimes):
        row += 1
        bg = _GREY_ROW if i % 2 == 0 else "FFFFFF"
        _cell(ws.cell(row, 1), d[0], bg=bg, align="center")
        _cell(ws.cell(row, 2), d[1], bg=bg, align="center")
        _cell(ws.cell(row, 3), d[2], bg=bg, align="center")
        _cell(ws.cell(row, 4), d[3], bg=bg)

    if not downtimes:
        row += 1
        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row, 1).value = "No downtime recorded."

    # ── Timestamp footer ────────────────────────────────────────────
    row += 2
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row, 1).value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.cell(row, 1).font  = Font(italic=True, color="999999", size=9)

    _autowidth(ws)


def _build_monthly_summary(wb, year: int, month: int, designer: str, monthly_rows):
    ws = wb.create_sheet("Monthly Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = f"Monthly Summary — {designer} — {year:04d}-{month:02d}"
    t.font  = Font(bold=True, size=13, color=_HEADER_FG)
    t.fill  = _header_fill(_TITLE_GREY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    row = 3
    headers = ["Date", "Reg Cases", "Cases (%)", "Downtime (min)", "Downtime (%)", "Total (%)", "Status"]
    for ci, h in enumerate(headers, 1):
        _hdr(ws.cell(row, ci), h)

    total_production_sum = 0.0
    for i, (d, n_cases, cases_pct, dt_min, dt_pct, total) in enumerate(monthly_rows):
        row += 1
        bg = _GREY_ROW if i % 2 == 0 else "FFFFFF"
        status = "✓ OK" if total >= 100 else ("▲ WARN" if total >= 95 else "✗ LOW")
        s_color = _GREEN if total >= 100 else (_YELLOW if total >= 95 else _RED)

        _cell(ws.cell(row, 1), d,                 bg=bg, align="center")
        _cell(ws.cell(row, 2), n_cases,            bg=bg, align="center")
        _cell(ws.cell(row, 3), f"{cases_pct:.2f}%",bg=bg, align="right")
        _cell(ws.cell(row, 4), f"{dt_min:.0f}",    bg=bg, align="center")
        _cell(ws.cell(row, 5), f"{dt_pct:.2f}%",   bg=bg, align="right")
        _cell(ws.cell(row, 6), f"{total:.2f}%",    bold=True, bg=_production_color(total),
              color=_HEADER_FG, align="center")
        _cell(ws.cell(row, 7), status, bg=s_color, color=_HEADER_FG, align="center")
        total_production_sum += total

    if monthly_rows:
        row += 1
        avg = total_production_sum / len(monthly_rows)
        ws.merge_cells(f"A{row}:E{row}")
        _cell(ws.cell(row, 1), "Monthly Average", bold=True, align="right", bg="EEEEEE")
        _cell(ws.cell(row, 6), f"{avg:.2f}%", bold=True,
              bg=_production_color(avg), color=_HEADER_FG, align="center")
        _cell(ws.cell(row, 7), "", bg="EEEEEE")

    _autowidth(ws)


# ── Team summary helpers ─────────────────────────────────────────────────────

def _safe_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no []:*?/\\ characters."""
    for ch in '[]:*?/\\':
        name = name.replace(ch, '-')
    return name[:31]


_DASH_SHEET = "Dashboard"


def _rebuild_dashboard(wb, today_str: str):
    """
    Rebuild (or create) the 'Dashboard' sheet as the first sheet.
    Shows one row per designer with today's data, a traffic-light color on
    Total%, a 'Last Sync' column, and a team-totals row at the bottom.
    Called every time any designer syncs so the supervisor always sees
    up-to-date data without opening individual sheets.
    """
    if _DASH_SHEET in wb.sheetnames:
        del wb[_DASH_SHEET]

    designer_sheets = [s for s in wb.sheetnames if s != _DASH_SHEET]

    ws = wb.create_sheet(_DASH_SHEET, 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = True

    refreshed = datetime.now().strftime("%H:%M")
    ws.merge_cells("A1:G1")
    title_cell = ws.cell(1, 1)
    title_cell.value = f"Production Dashboard  —  {today_str}    (refreshed {refreshed})"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=_TITLE_GREY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    col_headers = ["Designer", "Cases (%)", "Downtime (%)", "Total (%)", "Cases", "Last Sync", "Status"]
    for ci, h in enumerate(col_headers, 1):
        _hdr(ws.cell(2, ci), h)
    ws.freeze_panes = "A3"

    rows_data = []
    for sname in sorted(designer_sheets):
        src = wb[sname]
        today_row = None
        for row_cells in src.iter_rows(min_row=2):
            if row_cells[0].value == today_str:
                today_row = row_cells
                break

        if today_row is not None:
            def _v(cell):
                val = cell.value
                if isinstance(val, str) and val.endswith("%"):
                    try:
                        return float(val.rstrip("%"))
                    except ValueError:
                        return 0.0
                return val if val is not None else 0

            cases_pct = _v(today_row[2])
            dt_pct = _v(today_row[3])
            total_pct = _v(today_row[4])
            cases = today_row[5].value or 0
            ot_cases = today_row[6].value or 0
            last_sync = today_row[7].value if len(today_row) > 7 else "—"
        else:
            cases_pct = dt_pct = total_pct = cases = ot_cases = 0
            last_sync = "—"

        rows_data.append((sname, cases_pct, dt_pct, total_pct, cases, ot_cases, last_sync))

    for ri, (designer, cases_pct, dt_pct, total_pct, cases, ot_cases, last_sync) in enumerate(rows_data):
        row = ri + 3
        bg = _GREY_ROW if ri % 2 == 0 else "FFFFFF"
        no_data = (cases == 0 and ot_cases == 0 and last_sync == "—")

        total_bg = _production_color(total_pct) if not no_data else "E0E0E0"
        total_fg = _HEADER_FG if not no_data else "888888"
        status_text = "—" if no_data else (
            "🟢 On Track" if total_pct >= 95 else "🟡 At Risk" if total_pct >= 85 else "🔴 Behind"
        )
        status_bg = (
            "E0E0E0" if no_data else
            "C8E6C9" if total_pct >= 95 else
            "FFF9C4" if total_pct >= 85 else "FFCDD2"
        )
        status_fg = "888888" if no_data else "000000"

        _cell(ws.cell(row, 1), designer, bold=True, bg=bg)
        _cell(ws.cell(row, 2), f"{cases_pct:.1f}%" if not no_data else "—", align="right", bg=bg)
        _cell(ws.cell(row, 3), f"{dt_pct:.1f}%" if not no_data else "—", align="right", bg=bg)
        _cell(ws.cell(row, 4), f"{total_pct:.1f}%" if not no_data else "—", bold=True, color=total_fg, bg=total_bg, align="center")
        _cell(ws.cell(row, 5), cases if not no_data else "—", align="center", bg=bg)
        _cell(ws.cell(row, 6), last_sync or "—", align="center", bg=bg)
        _cell(ws.cell(row, 7), status_text, bold=True, color=status_fg, bg=status_bg, align="center")

    active = [r for r in rows_data if not (r[4] == 0 and r[5] == 0 and r[6] == "—")]
    if active:
        total_row = len(rows_data) + 3
        n = len(active)
        avg_cases = sum(r[1] for r in active) / n
        avg_dt = sum(r[2] for r in active) / n
        avg_total = sum(r[3] for r in active) / n
        sum_cases = sum(r[4] for r in active)
        TOTALS_BG = "2D2D2D"

        _hdr(ws.cell(total_row, 1), f"TEAM AVG  ({n} designers)", bg=TOTALS_BG, color=_HEADER_FG)
        _hdr(ws.cell(total_row, 2), f"{avg_cases:.1f}%", bg=TOTALS_BG, color=_HEADER_FG)
        _hdr(ws.cell(total_row, 3), f"{avg_dt:.1f}%", bg=TOTALS_BG, color=_HEADER_FG)
        _hdr(ws.cell(total_row, 4), f"{avg_total:.1f}%", bg=_production_color(avg_total), color=_HEADER_FG)
        _hdr(ws.cell(total_row, 5), sum_cases, bg=TOTALS_BG, color=_HEADER_FG)
        _hdr(ws.cell(total_row, 6), "", bg=TOTALS_BG, color=_HEADER_FG)
        _hdr(ws.cell(total_row, 7), "", bg=TOTALS_BG, color=_HEADER_FG)

    _autowidth(ws, extra=6)


def _rebuild_dashboard_file(productions_dir: str, today_str: str,
                            skip_live: bool = False,
                            skip_cache_check: bool = False,
                            force_snapshot: bool = False):
    """
        Read ALL _Summary_*.xlsx files and write _Dashboard.xlsx with two sheets:
            - "Dashboard": snapshot for today_str, one row per designer
            - "History": flat table for all dates/designers

        Skips the rebuild if no _Summary_*.xlsx file has changed since the
        last successful rebuild (mtime-based cache).

        Flags for bulk/historical operations:
            skip_live           → do not touch the live _Dashboard.xlsx.
            skip_cache_check    → ignore the mtime cache (force build).
            force_snapshot      → overwrite existing dated snapshot.
    """
    global _last_dashboard_max_mtime
    import glob as _glob
    from openpyxl.utils import get_column_letter as _gcl

    # Look in per-designer folders first, fall back to old flat location
    summary_files = sorted(
        _glob.glob(os.path.join(productions_dir, "*", "_Summary.xlsx"))
    )
    # Also pick up any old-style summaries not yet migrated
    summary_files += sorted(
        _glob.glob(os.path.join(productions_dir, "_Summary_*.xlsx"))
    )

    # ── Skip rebuild if no summary file changed since last time ──────────────
    dashboard_path = os.path.join(productions_dir, "_Dashboard.xlsx")
    if not skip_cache_check and summary_files:
        max_mtime = max(os.path.getmtime(f) for f in summary_files)
        if max_mtime <= _last_dashboard_max_mtime and os.path.exists(dashboard_path):
            print("[sharepoint_sync] No summary files changed, skipping Dashboard rebuild.")
            return

    def _pct(v):
        if isinstance(v, str) and v.endswith("%"):
            try:
                return float(v.rstrip("%"))
            except ValueError:
                return 0.0
        return float(v) if v is not None else 0.0

    def _num(row, idx, default=0):
        try:
            v = row[idx]
            return int(v) if v is not None else default
        except (IndexError, TypeError, ValueError):
            return default

    def _flt(row, idx, default=0.0):
        try:
            v = row[idx]
            return float(v) if v is not None else default
        except (IndexError, TypeError, ValueError):
            return default

    # ── Load all data from every summary file ────────────────────────────────
    # Entry tuple indices:
    #  0 designer 1 date 2 week 3 cases_pct 4 dt_pct 5 total_pct
    #  6 reg_cases 7 ot_cases 8 ue
    #  9 reg_explicit(dict) 10 reg_other
    #  11 ot_explicit(dict) 12 ot_other
    #  13 last_sync
    all_rows = []
    today_data = {}   # designer → entry tuple for today_str

    def _parse_summary_type_counts(row_values):
        reg_explicit = {name: 0 for name in CASE_TYPE_COLUMNS}
        ot_explicit = {name: 0 for name in CASE_TYPE_COLUMNS}
        reg_other = 0
        ot_other = 0
        last_sync = "—"

        # Historic case-type sets (each added a column when new types shipped).
        TYPES_10 = CASE_TYPE_COLUMNS  # includes "New Impressions"
        TYPES_9 = [t for t in CASE_TYPE_COLUMNS if t != "New Impressions"]

        # Newest schema (32 cols, 10 case types incl. New Impressions)
        if len(row_values) >= 32:
            for i, case_type in enumerate(TYPES_10):
                reg_explicit[case_type] = _num(row_values, 9 + i)
            reg_other = _num(row_values, 19)
            for i, case_type in enumerate(TYPES_10):
                ot_explicit[case_type] = _num(row_values, 20 + i)
            ot_other = _num(row_values, 30)
            last_sync = row_values[31] if row_values[31] else "—"
            return reg_explicit, reg_other, ot_explicit, ot_other, last_sync

        # Previous 30-col schema (9 case types, no New Impressions)
        if len(row_values) >= 30:
            for i, case_type in enumerate(TYPES_9):
                reg_explicit[case_type] = _num(row_values, 9 + i)
            reg_other = _num(row_values, 18)
            for i, case_type in enumerate(TYPES_9):
                ot_explicit[case_type] = _num(row_values, 19 + i)
            ot_other = _num(row_values, 28)
            last_sync = row_values[29] if row_values[29] else "—"
            return reg_explicit, reg_other, ot_explicit, ot_other, last_sync

        # Previous 29-col schema (9 case types, one less leading column)
        if len(row_values) >= 29:
            for i, case_type in enumerate(TYPES_9):
                reg_explicit[case_type] = _num(row_values, 8 + i)
            reg_other = _num(row_values, 17)
            for i, case_type in enumerate(TYPES_9):
                ot_explicit[case_type] = _num(row_values, 18 + i)
            ot_other = _num(row_values, 27)
            last_sync = row_values[28] if row_values[28] else "—"
            return reg_explicit, reg_other, ot_explicit, ot_other, last_sync

        # Previous bucket schema (21 cols)
        if len(row_values) >= 21:
            reg_explicit["Primary"] = _num(row_values, 8)
            reg_explicit["Secondary"] = _num(row_values, 9)
            reg_explicit["CR"] = _num(row_values, 10)
            reg_other = _num(row_values, 11) + _num(row_values, 12) + _num(row_values, 13)

            ot_explicit["Primary"] = _num(row_values, 14)
            ot_explicit["Secondary"] = _num(row_values, 15)
            ot_explicit["CR"] = _num(row_values, 16)
            ot_other = _num(row_values, 17) + _num(row_values, 18) + _num(row_values, 19)

            last_sync = row_values[20] if row_values[20] else "—"
            return reg_explicit, reg_other, ot_explicit, ot_other, last_sync

        # Old schema (15 cols)
        reg_explicit["Primary"] = _num(row_values, 8)
        reg_explicit["Secondary"] = _num(row_values, 9)
        reg_explicit["CR"] = _num(row_values, 10)
        reg_other = _num(row_values, 11)
        ot_explicit["Primary"] = _num(row_values, 12)
        ot_other = _num(row_values, 13)
        last_sync = row_values[14] if len(row_values) > 14 and row_values[14] else "—"
        return reg_explicit, reg_other, ot_explicit, ot_other, last_sync

    for sf in summary_files:
        try:
            swb = openpyxl.load_workbook(sf, read_only=True, data_only=True)
            sws = swb.active
            base = os.path.basename(sf)
            if base == "_Summary.xlsx":
                # New format: Productions/<DesignerName>/_Summary.xlsx
                designer_name = os.path.basename(os.path.dirname(sf)).replace("_", " ")
            else:
                # Old format: Productions/_Summary_<DesignerName>.xlsx
                designer_name = base[len("_Summary_"):-len(".xlsx")].replace("_", " ")
            for row in sws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                d_str      = str(row[0])
                week       = row[1] or ""
                cases_pct  = _pct(row[2])
                dt_pct     = _pct(row[3])
                total_pct  = _pct(row[4])
                reg_cases  = _num(row, 5)
                ot_cases   = _num(row, 6)
                ue         = _flt(row, 7)
                ue_cases   = _flt(row, 8) if len(row) >= 30 else 0.0
                reg_explicit, reg_other, ot_explicit, ot_other, last_sync = _parse_summary_type_counts(row)

                entry = (designer_name, d_str, week, cases_pct, dt_pct,
                         total_pct, reg_cases, ot_cases, ue,
                         ue_cases,
                         reg_explicit, reg_other,
                         ot_explicit, ot_other,
                         last_sync)
                all_rows.append(entry)
                if d_str == today_str:
                    today_data[designer_name] = entry
            swb.close()
        except Exception as exc:
            log_event("sharepoint_sync", f"dashboard source summary parse skipped for {sf}: {exc}", level="WARN")
            continue

    # Build snapshot: include all known designers (blank row if no data today)
    all_designers = sorted({r[0] for r in all_rows})
    _blank = lambda d: (d, today_str, "", 0.0, 0.0, 0.0,
                        0, 0, 0.0,
                        0.0,
                        {name: 0 for name in CASE_TYPE_COLUMNS}, 0,
                        {name: 0 for name in CASE_TYPE_COLUMNS}, 0,
                        "—")
    today_rows = [today_data.get(d) or _blank(d) for d in all_designers]

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Dashboard (today_str snapshot)
    # ════════════════════════════════════════════════════════════════════════
    # Columns: Designer | Cases% | DT% | Total% | UE(Cases) | UE(Total) | Reg |
    #          Reg <all CASE_TYPE_COLUMNS> | Reg Other |
    #          Last Sync | Status
    dash_headers = [
        "Designer", "Cases (%)", "Downtime (%)", "Total (%)",
        "UE (Cases)", "UE (w/ DT)", "Reg Cases",
    ] + [f"Reg {name}" for name in CASE_TYPE_COLUMNS] + [
        "Reg Other", "Last Sync", "Status"
    ]
    DASH_COLS = len(dash_headers)
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_view.showGridLines = False

    refreshed = datetime.now().strftime("%H:%M")
    ws_dash.merge_cells(f"A1:{_gcl(DASH_COLS)}1")
    tc = ws_dash.cell(1, 1)
    tc.value = (f"Production Dashboard  —  {today_str}"
                f"    (refreshed {refreshed})")
    tc.font      = Font(bold=True, size=13, color="FFFFFF")
    tc.fill      = PatternFill("solid", fgColor=_TITLE_GREY)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 28

    # ── Row 2: live-data notice ──────────────────────────────────────────
    # The xlsx is rebuilt each time someone saves a case, but OneDrive has
    # eventual consistency, so this snapshot is generally a few minutes
    # behind the source data. The app's Dashboard → Equipo tab reads the
    # same source files every 30 s so it's always the freshest view.
    ws_dash.merge_cells(f"A2:{_gcl(DASH_COLS)}2")
    nc = ws_dash.cell(2, 1)
    nc.value = ("Para datos en tiempo real abre la app  →  Dashboard  →  Equipo. "
                "Este archivo se actualiza solo cuando alguien guarda un caso.")
    nc.font      = Font(italic=True, size=10, color="555555")
    nc.fill      = PatternFill("solid", fgColor="FFF8E1")
    nc.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[2].height = 22

    for ci, h in enumerate(dash_headers, 1):
        _hdr(ws_dash.cell(3, ci), h)
    ws_dash.freeze_panes = "A4"

    for ri, entry in enumerate(today_rows):
        (
            designer_name,
            _d,
            _w,
            cases_pct,
            dt_pct,
            total_pct,
            reg_cases,
            ot_cases,
            ue,
            ue_cases,
            reg_explicit,
            reg_other,
            ot_explicit,
            ot_other,
            last_sync,
        ) = entry
        row = ri + 4  # row 1 = title, row 2 = live-data notice, row 3 = headers
        bg = _GREY_ROW if ri % 2 == 0 else "FFFFFF"
        no_data = (reg_cases == 0 and ot_cases == 0 and last_sync == "—")

        total_bg = _production_color(total_pct) if not no_data else "E0E0E0"
        total_fg = _HEADER_FG if not no_data else "888888"
        status_text = (
            "—" if no_data else
            "On Track" if total_pct >= 95 else
            "At Risk" if total_pct >= 85 else "Behind"
        )
        status_bg = (
            "E0E0E0" if no_data else
            "C8E6C9" if total_pct >= 95 else
            "FFF9C4" if total_pct >= 85 else "FFCDD2"
        )
        status_fg = "888888" if no_data else "000000"
        dash = "—"

        _cell(ws_dash.cell(row, 1), designer_name, bold=True, bg=bg)
        _cell(ws_dash.cell(row, 2), f"{cases_pct:.1f}%" if not no_data else dash, align="right", bg=bg)
        _cell(ws_dash.cell(row, 3), f"{dt_pct:.1f}%" if not no_data else dash, align="right", bg=bg)
        _cell(
            ws_dash.cell(row, 4),
            f"{total_pct:.1f}%" if not no_data else dash,
            bold=True,
            color=total_fg,
            bg=total_bg,
            align="center",
        )
        _cell(ws_dash.cell(row, 5), f"{ue_cases:.2f}" if not no_data else dash, align="right", bg=bg)
        _cell(ws_dash.cell(row, 6), f"{ue:.2f}" if not no_data else dash, align="right", bg=bg)
        _cell(ws_dash.cell(row, 7), reg_cases if not no_data else dash, align="center", bg=bg)

        write_col = 8
        for case_type in CASE_TYPE_COLUMNS:
            _cell(
                ws_dash.cell(row, write_col),
                reg_explicit.get(case_type, 0) if not no_data else dash,
                align="center",
                bg=bg,
            )
            write_col += 1

        _cell(ws_dash.cell(row, write_col), reg_other if not no_data else dash, align="center", bg=bg)
        write_col += 1
        _cell(ws_dash.cell(row, write_col), last_sync or dash, align="center", bg=bg)
        write_col += 1
        _cell(ws_dash.cell(row, write_col), status_text, bold=True, color=status_fg, bg=status_bg, align="center")

    # Team averages row
    active_today = [e for e in today_rows if not (e[6] == 0 and e[7] == 0 and e[14] == "—")]
    if active_today:
        tr = len(today_rows) + 4  # +4 to account for title/notice/header
        n  = len(active_today)
        TOTALS_BG = "2D2D2D"
        _hdr(ws_dash.cell(tr,  1), f"TEAM AVG  ({n} designers)",
             bg=TOTALS_BG, color=_HEADER_FG)
        _hdr(ws_dash.cell(tr,  2),
             f"{sum(e[3] for e in active_today)/n:.1f}%",
             bg=TOTALS_BG, color=_HEADER_FG)
        _hdr(ws_dash.cell(tr,  3),
             f"{sum(e[4] for e in active_today)/n:.1f}%",
             bg=TOTALS_BG, color=_HEADER_FG)
        avg_total = sum(e[5] for e in active_today) / n
        _hdr(ws_dash.cell(tr,  4), f"{avg_total:.1f}%",
             bg=_production_color(avg_total), color=_HEADER_FG)
        _hdr(ws_dash.cell(tr,  5),
             f"{sum(e[9]  for e in active_today):.2f}",
             bg=TOTALS_BG, color=_HEADER_FG)
        _hdr(ws_dash.cell(tr,  6),
             f"{sum(e[8]  for e in active_today):.2f}",
             bg=TOTALS_BG, color=_HEADER_FG)
        _hdr(ws_dash.cell(tr,  7), sum(e[6]  for e in active_today),
             bg=TOTALS_BG, color=_HEADER_FG)
        write_col = 8
        for case_type in CASE_TYPE_COLUMNS:
            _hdr(ws_dash.cell(tr, write_col),
                 sum(e[10].get(case_type, 0) for e in active_today),
                 bg=TOTALS_BG, color=_HEADER_FG)
            write_col += 1
        _hdr(ws_dash.cell(tr, write_col), sum(e[11] for e in active_today),
             bg=TOTALS_BG, color=_HEADER_FG)
        write_col += 1
        _hdr(ws_dash.cell(tr, write_col), "", bg=TOTALS_BG, color=_HEADER_FG)
        write_col += 1
        _hdr(ws_dash.cell(tr, write_col), "", bg=TOTALS_BG, color=_HEADER_FG)

    _autowidth(ws_dash, extra=1)
    for ci in range(1, DASH_COLS + 1):
        col = _gcl(ci)
        if ci == 1:
            ws_dash.column_dimensions[col].width = 18
        elif ci in (2, 3, 4):
            ws_dash.column_dimensions[col].width = 11
        elif ci in (5, 6):
            ws_dash.column_dimensions[col].width = 10
        elif ci == 7:
            ws_dash.column_dimensions[col].width = 9
        elif ci in (DASH_COLS - 1, DASH_COLS):
            ws_dash.column_dimensions[col].width = 10
        else:
            ws_dash.column_dimensions[col].width = min(ws_dash.column_dimensions[col].width, 12)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2 — History (all dates × all designers, filterable table)
    # Layout:
    #   Row 1  — date-picker header  (label | date input cell | instruction)
    #   Row 2  — thin spacer
    #   Row 3  — table column headers
    #   Row 4+ — data rows
    # ════════════════════════════════════════════════════════════════════════
    import datetime as _dt

    hist_headers = [
        "Date", "Designer", "Cases (%)", "Downtime (%)", "Total (%)",
        "UE (Cases)", "UE (w/ DT)",
        "Reg Cases",
    ] + [f"Reg {name}" for name in CASE_TYPE_COLUMNS] + [
        "Reg Other", "Status"
    ]
    HIST_COLS = len(hist_headers)
    HDR_ROW   = 3   # table header row
    DATA_ROW  = 4   # first data row

    ws_hist = wb.create_sheet("History")
    ws_hist.sheet_view.showGridLines = False

    # ── Row 1: date-picker ───────────────────────────────────────────────
    try:
        _filter_date = _dt.datetime.strptime(today_str, "%Y-%m-%d").date()
    except Exception as exc:
        log_event("sharepoint_sync", f"history filter date parse failed for {today_str}: {exc}", level="WARN")
        _filter_date = _dt.date.today()

    # A1 — label
    lbl = ws_hist.cell(1, 1)
    lbl.value     = "View date:"
    lbl.font      = Font(bold=True, color="FFFFFF", size=11)
    lbl.fill      = PatternFill("solid", fgColor="2E75B6")
    lbl.alignment = Alignment(horizontal="right", vertical="center")

    # B1 — date input cell (actual date value so Excel treats it as a date)
    dc = ws_hist.cell(1, 2)
    dc.value         = _filter_date
    dc.number_format = "YYYY-MM-DD"
    dc.font          = Font(bold=True, size=12, color="1F3864")
    dc.fill          = PatternFill("solid", fgColor="D9E1F2")
    dc.alignment     = Alignment(horizontal="center", vertical="center")
    _med = Side(style="medium", color="2E75B6")
    dc.border = Border(left=_med, right=_med, top=_med, bottom=_med)

    # DataValidation: accept any date (serial > 0)
    dv = DataValidation(
        type="date", operator="greaterThan", formula1="0",
        showErrorMessage=False, showInputMessage=True,
        promptTitle="Date filter",
        prompt="Edit this date, then use the Date ▼ dropdown on the table to filter"
    )
    ws_hist.add_data_validation(dv)
    dv.add(dc)

    # C1:L1 — instruction text
    ws_hist.merge_cells(f"C1:{_gcl(HIST_COLS)}1")
    inst = ws_hist.cell(1, 3)
    inst.value     = ("← Edit this date, then click the  Date ▼  "
                      "dropdown on the table header to filter")
    inst.font      = Font(italic=True, color="888888", size=10)
    inst.fill      = PatternFill("solid", fgColor="F2F2F2")
    inst.alignment = Alignment(horizontal="left", vertical="center")
    ws_hist.row_dimensions[1].height = 26

    # ── Row 2: thin spacer ───────────────────────────────────────────────
    for ci in range(1, HIST_COLS + 1):
        ws_hist.cell(2, ci).fill = PatternFill("solid", fgColor="E8EEF7")
    ws_hist.row_dimensions[2].height = 4

    # ── Row 3: column headers ────────────────────────────────────────────
    for ci, h in enumerate(hist_headers, 1):
        _hdr(ws_hist.cell(HDR_ROW, ci), h)
    ws_hist.freeze_panes = f"A{DATA_ROW}"

    # ── Rows 4+: data ────────────────────────────────────────────────────
    # Sort: date desc, then designer asc
    sorted_rows = sorted(all_rows, key=lambda r: (r[1], r[0]))
    sorted_rows.sort(key=lambda r: r[1], reverse=True)

    for ri, entry in enumerate(sorted_rows):
        (designer_name, d_str, _w, cases_pct, dt_pct, total_pct,
         reg_cases, ot_cases, ue,
         ue_cases,
         reg_explicit, reg_other,
         ot_explicit, ot_other,
         last_sync) = entry
        row      = ri + DATA_ROW
        bg       = _GREY_ROW if ri % 2 == 0 else "FFFFFF"
        total_bg = _production_color(total_pct)
        status_text = ("On Track" if total_pct >= 95 else
                       "At Risk"  if total_pct >= 85 else "Behind")
        status_bg   = ("C8E6C9" if total_pct >= 95 else
                       "FFF9C4" if total_pct >= 85 else "FFCDD2")

        _cell(ws_hist.cell(row,  1), d_str,              align="center", bg=bg)
        _cell(ws_hist.cell(row,  2), designer_name,       bold=True,     bg=bg)
        _cell(ws_hist.cell(row,  3), f"{cases_pct:.1f}%", align="right", bg=bg)
        _cell(ws_hist.cell(row,  4), f"{dt_pct:.1f}%",    align="right", bg=bg)
        _cell(ws_hist.cell(row,  5), f"{total_pct:.1f}%",
              bold=True, color=_HEADER_FG, bg=total_bg, align="center")
        _cell(ws_hist.cell(row,  6), f"{ue_cases:.2f}",   align="right", bg=bg)
        _cell(ws_hist.cell(row,  7), f"{ue:.2f}",          align="right", bg=bg)
        _cell(ws_hist.cell(row,  8), reg_cases, align="center", bg=bg)

        write_col = 9
        for case_type in CASE_TYPE_COLUMNS:
            _cell(
                ws_hist.cell(row, write_col),
                reg_explicit.get(case_type, 0),
                align="center",
                bg=bg,
            )
            write_col += 1

        _cell(ws_hist.cell(row, write_col), reg_other, align="center", bg=bg)
        write_col += 1
        _cell(ws_hist.cell(row, write_col), status_text, bold=True, bg=status_bg, align="center")

    # ── Excel Table + pre-applied date filter ────────────────────────────
    if sorted_rows:
        last_row = len(sorted_rows) + HDR_ROW
        ref = f"A{HDR_ROW}:{_gcl(HIST_COLS)}{last_row}"
        tbl = Table(displayName="HistoryTable", ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True,   showColumnStripes=False)
        ws_hist.add_table(tbl)

        # Pre-apply filter on Date column (colId=0) to today_str
        # so the file opens already showing only the relevant day.
        ws_hist.auto_filter.ref = ref
        fc = FilterColumn(colId=0)
        fc.filters = Filters(filter=[today_str])
        ws_hist.auto_filter.filterColumn = [fc]

    _autowidth(ws_hist, extra=1)
    for ci in range(1, HIST_COLS + 1):
        col = _gcl(ci)
        if ci in (1, 2):
            ws_hist.column_dimensions[col].width = 13
        elif ci in (3, 4, 5):
            ws_hist.column_dimensions[col].width = 11
        elif ci in (6, 7):
            ws_hist.column_dimensions[col].width = 10
        elif ci == 8:
            ws_hist.column_dimensions[col].width = 9
        elif ci == HIST_COLS:
            ws_hist.column_dimensions[col].width = 10
        else:
            ws_hist.column_dimensions[col].width = min(ws_hist.column_dimensions[col].width, 12)

    dashboard_path = os.path.join(productions_dir, "_Dashboard.xlsx")
    try:
        if not skip_live:
            try:
                os.remove(dashboard_path)
            except FileNotFoundError:
                pass
            _save_dashboard_verified(wb, dashboard_path)
            # Update mtime cache so next call skips rebuild if nothing changed
            if summary_files:
                _last_dashboard_max_mtime = max(os.path.getmtime(f) for f in summary_files if os.path.exists(f))
    except PermissionError:
        # If dashboard is open/locked (Excel, preview, etc.), skip silently.
        # Daily and summary files were already saved.
        return

    # Save a dated snapshot. We ALWAYS overwrite — previously this skipped
    # the write when the file already existed, which froze each daily
    # snapshot at whatever state the *first* designer's sync produced
    # (typically near-empty, before most of the team had logged anything).
    # Late syncs and backdated edits then never propagated to the snapshot,
    # making historical dashboards look like nobody worked that day.
    snapshots_dir = os.path.join(productions_dir, "Dashboards")
    os.makedirs(snapshots_dir, exist_ok=True)
    snapshot_path = os.path.join(snapshots_dir, f"_Dashboard_{today_str}.xlsx")
    try:
        _save_atomic(wb, snapshot_path)
    except PermissionError:
        # Someone has the historical snapshot open; skip silently — next
        # rebuild will retry.
        pass

    # Clean up OneDrive conflict copies (e.g. _Dashboard-CRI-MACHINE.xlsx)
    if not skip_live:
        _cleanup_onedrive_conflicts(productions_dir)


def _update_team_summary(productions_dir: str, designer: str, target_date: str,
                         total_cases_pct: float, total_downtime_min: float,
                         n_cases: int, n_ot_cases: int,
                         ue_total: float = 0.0,
                         reg_type_counts: dict | None = None,
                         ot_type_counts:  dict | None = None,
                         ue_cases: float = 0.0,
                         rebuild_dashboard: bool = True):
    """
    NEW ARCHITECTURE — no shared-file lock problem:

    Step 1: Write _Summary_<DesignerName>.xlsx  (only THIS designer ever writes it)
            → no conflicts possible, always succeeds even if Dashboard is open.

    Step 2: Rebuild _Dashboard.xlsx by reading ALL _Summary_*.xlsx files.
            → if Excel has _Dashboard.xlsx open (locked), skip silently;
              it will be rebuilt on the next sync.
    """
    if reg_type_counts is None:
        reg_type_counts = {}
    if ot_type_counts is None:
        ot_type_counts = {}
    safe_name = designer.replace(" ", "_").replace("/", "-")

    # Per-designer folder: Productions/<DesignerName>/
    designer_dir = os.path.join(productions_dir, safe_name)
    os.makedirs(designer_dir, exist_ok=True)
    summary_file = os.path.join(designer_dir, "_Summary.xlsx")

    # Migration: move old summary from Productions/_Summary_<name>.xlsx to new location
    old_summary = os.path.join(productions_dir, f"_Summary_{safe_name}.xlsx")
    if os.path.exists(old_summary) and not os.path.exists(summary_file):
        try:
            import shutil
            shutil.move(old_summary, summary_file)
            print(f"[sharepoint_sync] Migrated summary to {designer_dir}")
        except Exception as exc:
            log_event("sharepoint_sync", f"summary migration failed {old_summary} -> {summary_file}: {exc}", level="WARN")

    d         = date.fromisoformat(target_date)
    week_num  = d.isocalendar()[1]
    dt_pct    = (total_downtime_min / DAILY_BASE_MINUTES) * 100
    total_pct = total_cases_pct + dt_pct

    os.makedirs(productions_dir, exist_ok=True)

    # ── Step 1: Read existing historical rows into memory (skip today) ────────
    # We read first, then delete, then write fresh — no in-place overwrite
    # so there is no file-lock issue.
    history_rows = []   # list of raw value tuples, date != target_date
    if os.path.exists(summary_file):
        try:
            old_wb = openpyxl.load_workbook(summary_file, read_only=True,
                                            data_only=True)
            old_ws = old_wb.active
            for row in old_ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[0] != target_date:
                    history_rows.append(row)
            old_wb.close()
        except Exception as _read_exc:
            # SAFETY: if we cannot read the existing file (locked by OneDrive,
            # open in Excel, or corrupt) we must NOT delete it — that would
            # permanently erase all historical rows.  Surface the error so the
            # caller knows the sync did not complete.
            raise PermissionError(
                f"Cannot read the existing summary file — it may be open in "
                f"Excel or locked by OneDrive.\n\n"
                f"File: {summary_file}\n\n"
                f"Close the file (and wait for OneDrive to finish syncing), "
                f"then try again.\n\nDetail: {_read_exc}"
            ) from _read_exc
        # Read succeeded — safe to delete and rewrite
        try:
            os.remove(summary_file)
        except Exception as exc:
            log_event("sharepoint_sync", f"could not remove existing summary before rewrite: {exc}", level="WARN")

    # ── Step 2: Build fresh workbook with all rows ────────────────────────────
    # Column layout (1-indexed):
    # 1:Date 2:Week 3:Cases% 4:DT% 5:Total% 6:RegCases 7:OTCases
    # 8:UE(Total) 9:UE(Cases)
    # 10..(9+N): Reg explicit CASE_TYPE_COLUMNS
    # 10+N:Reg Other
    # (11+N)..(10+2N): OT explicit CASE_TYPE_COLUMNS
    # 11+2N:OT Other
    # 12+2N:LastSync
    NCOLS = 12 + 2 * len(CASE_TYPE_COLUMNS)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = safe_name
    ws.sheet_view.showGridLines = False
    col_headers = [
        "Date", "Week", "Cases (%)", "Downtime (%)", "Total (%)",
        "Reg Cases", "OT Cases", "UE (Total)", "UE (Cases)",
    ] + [f"Reg {name}" for name in CASE_TYPE_COLUMNS] + [
        "Reg Other",
    ] + [f"OT {name}" for name in CASE_TYPE_COLUMNS] + [
        "OT Other", "Last Sync"
    ]
    for ci, h in enumerate(col_headers, 1):
        _hdr(ws.cell(1, ci), h)
    ws.freeze_panes = "A2"

    # Re-write history rows sorted oldest first (pad/trim to NCOLS if needed)
    history_rows.sort(key=lambda r: r[0])
    for ri, hrow in enumerate(history_rows):
        row_i = ri + 2
        bg = _GREY_ROW if ri % 2 == 0 else "FFFFFF"
        padded = list(hrow) + [None] * NCOLS
        for ci in range(1, NCOLS + 1):
            _cell(ws.cell(row_i, ci), padded[ci - 1],
                  align="center" if ci in (1, 2, 6, 7, 15) else "right",
                  bg=bg)

    # Compute today's explicit type counts (REG + OT)
    reg_explicit, reg_other = _bucket_case_types(reg_type_counts)
    ot_explicit, ot_other = _bucket_case_types(ot_type_counts)

    # Write today's fresh data
    today_ri  = len(history_rows)
    write_row = today_ri + 2
    bg = _GREY_ROW if today_ri % 2 == 0 else "FFFFFF"

    _cell(ws.cell(write_row,  1), target_date,               align="center", bg=bg)
    _cell(ws.cell(write_row,  2), f"W{week_num:02d}",        align="center", bg=bg)
    _cell(ws.cell(write_row,  3), f"{total_cases_pct:.2f}%", align="right",  bg=bg)
    _cell(ws.cell(write_row,  4), f"{dt_pct:.2f}%",          align="right",  bg=bg)
    _cell(ws.cell(write_row,  5), f"{total_pct:.2f}%",       bold=True,
          bg=_production_color(total_pct), color=_HEADER_FG, align="center")
    _cell(ws.cell(write_row,  6), n_cases,                   align="center", bg=bg)
    _cell(ws.cell(write_row,  7), n_ot_cases,                align="center", bg=bg)
    _cell(ws.cell(write_row,  8), round(ue_total, 2),        align="right",  bg=bg)
    _cell(ws.cell(write_row,  9), round(ue_cases, 2),        align="right",  bg=bg)
    write_col = 10
    for case_type in CASE_TYPE_COLUMNS:
        _cell(ws.cell(write_row, write_col), reg_explicit.get(case_type, 0), align="center", bg=bg)
        write_col += 1
    _cell(ws.cell(write_row, write_col), reg_other, align="center", bg=bg)
    write_col += 1
    for case_type in CASE_TYPE_COLUMNS:
        _cell(ws.cell(write_row, write_col), ot_explicit.get(case_type, 0), align="center", bg=bg)
        write_col += 1
    _cell(ws.cell(write_row, write_col), ot_other, align="center", bg=bg)
    write_col += 1
    _cell(ws.cell(write_row, write_col), datetime.now().strftime("%H:%M"), align="center", bg=bg)

    _autowidth(ws)
    wb.save(summary_file)   # fresh file, no lock possible

    # ── Step 2: Rebuild _Dashboard.xlsx from all _Summary_*.xlsx ─────────────
    if rebuild_dashboard:
        _rebuild_dashboard_file(productions_dir, target_date)


def _save_dashboard_verified(wb, final_path: str, verify_delay: float = 5.0, retries: int = 2):
    """Save Dashboard, re-hash after delay. If OneDrive mutated the file
    (conflict merge or another designer's save landed), re-save so our
    version wins. Max `retries` re-saves.
    """
    import hashlib

    def _sha(path: str):
        try:
            with open(path, "rb") as fh:
                return hashlib.sha256(fh.read()).hexdigest()
        except Exception:
            return None

    _save_atomic(wb, final_path)
    original = _sha(final_path)
    if original is None:
        return

    for attempt in range(retries):
        time.sleep(verify_delay)
        current = _sha(final_path)
        if current == original:
            return
        log_event("sharepoint_sync",
                  f"dashboard mutated after save (hash mismatch), re-saving attempt {attempt + 1}",
                  level="WARN")
        try:
            _save_atomic(wb, final_path)
        except Exception as exc:
            log_event("sharepoint_sync", f"dashboard re-save failed: {exc}", level="WARN")
            return
        original = _sha(final_path)
        if original is None:
            return


def _save_atomic(wb, final_path: str, retries: int = 8):
    """Save workbook directly to final_path, retrying with random jitter if
    OneDrive or another user holds a lock.

    Random delay between retries prevents 25 designers from all hammering the
    file at the same millisecond after a failed attempt.

    We do NOT use a temp-file + os.replace() because that swaps the file inode,
    which OneDrive interprets as two conflicting versions of the same file.
    Instead we overwrite in-place with retries so the inode stays the same and
    OneDrive just sees an updated file with no conflict.
    """
    import time, random
    os.makedirs(os.path.dirname(final_path) or ".", exist_ok=True)
    last_exc: Exception = RuntimeError("no attempts made")
    for attempt in range(max(1, retries)):
        try:
            wb.save(final_path)
            return                      # success
        except PermissionError as exc:
            last_exc = exc
            if attempt < retries - 1:
                # Random jitter 1–6 s so each designer backs off independently
                time.sleep(random.uniform(*_SP_UPLOAD_JITTER))
        except Exception:
            raise
    raise last_exc


# ── Public entry point ───────────────────────────────────────────────────────

def export_to_sharepoint(target_date: str | None = None) -> tuple[bool, str]:
    """
    Generate the Excel report and save to the configured export folder.

    Returns (success: bool, message: str)
    """
    if not _OPENPYXL_OK:
        return False, f"openpyxl load error: {_OPENPYXL_ERROR}"

    cfg = load_config()
    designer     = cfg.get("designer_name", "Designer").strip() or "Designer"
    export_folder = cfg.get("export_folder", "").strip()

    if not export_folder:
        return False, "Export folder not configured. Please set it in Settings."
    if not os.path.isdir(export_folder):
        return False, f"Export folder not found:\n{export_folder}"

    if target_date is None:
        target_date = date.today().isoformat()

    # Gather data
    cases, ot_cases, downtimes, total_cases_pct, total_downtime_min = \
        _get_daily_data(target_date)
    year  = int(target_date[:4])
    month = int(target_date[5:7])
    monthly_rows = _get_monthly_data(year, month)

    # ── Build structured folder path ─────────────────────────────────────────
    # Productions/
    #   2026-02/
    #     Week-08/
    #       2026-02-23/
    #         Gerardo_Production_2026-02-23.xlsx
    d         = date.fromisoformat(target_date)
    week_num  = d.isocalendar()[1]
    month_str = d.strftime("%Y-%m")          # "2026-02"
    week_str  = f"Week-{week_num:02d}"        # "Week-08"

    productions_dir = os.path.join(export_folder, "Productions")
    day_dir = os.path.join(productions_dir, month_str, week_str, target_date)
    os.makedirs(day_dir, exist_ok=True)

    # Compute UE upfront so daily file can show it too
    ue_cases, ue_total, reg_type_counts, ot_type_counts = _get_ue_and_types(target_date)

    # Build workbook
    wb = openpyxl.Workbook()
    _build_daily_summary(wb, target_date, designer,
                         cases, ot_cases, downtimes,
                         total_cases_pct, total_downtime_min,
                         ue_cases=ue_cases, ue_total=ue_total)
    _build_monthly_summary(wb, year, month, designer, monthly_rows)

    # Save individual daily file
    safe_name = designer.replace(" ", "_").replace("/", "-")
    filename  = f"{safe_name}_Production_{target_date}.xlsx"
    out_path  = os.path.join(day_dir, filename)

    try:
        _save_atomic(wb, out_path)
    except Exception as e:
        return False, f"Could not save daily file:\n{e}"

    # ── Update shared team summary ────────────────────────────────────────────
    try:
        _update_team_summary(
            productions_dir, designer, target_date,
            total_cases_pct, total_downtime_min,
            len(cases), len(ot_cases),
            ue_total, reg_type_counts, ot_type_counts,
            ue_cases=ue_cases,
        )
    except Exception as e:
        # Return False so the caller knows something went wrong
        return False, f"Daily report saved but team summary failed:\n{e}"

    safe_name = designer.replace(" ", "_").replace("/", "-")
    return True, (
        f"Report saved:\n{out_path}"
        f"\n\nSummary updated:\n{productions_dir}\\{safe_name}\\_Summary.xlsx"
        f"\n\nDashboard rebuilt:\n{productions_dir}\\_Dashboard.xlsx"
        f"\n\nOneDrive will sync to SharePoint automatically."
    )


def export_all_missing_to_sharepoint(progress_cb=None) -> tuple[bool, str]:
    """Bulk upload daily files for every date with cases in the local DB that
    is missing or corrupt in the shared folder. Valid existing files are
    skipped (never overwritten). The team summary is updated per date and the
    Dashboard is rebuilt only once at the end (protected by the stagger +
    cooldown + verify pipeline).

    progress_cb(index, total, message): optional callable for UI feedback.
    """
    if not _OPENPYXL_OK:
        return False, f"openpyxl load error: {_OPENPYXL_ERROR}"

    cfg = load_config()
    designer = (cfg.get("designer_name", "Designer") or "Designer").strip() or "Designer"
    export_folder = (cfg.get("export_folder", "") or "").strip()
    if not export_folder:
        return False, "Export folder not configured. Please set it in Settings."
    if not os.path.isdir(export_folder):
        return False, f"Export folder not found:\n{export_folder}"

    productions_dir = os.path.join(export_folder, "Productions")
    safe_name = designer.replace(" ", "_").replace("/", "-")

    # ── Collect every date with local data ───────────────────────────────────
    conn = _db()
    cur = conn.cursor()
    cur.execute(
        "SELECT DISTINCT fecha FROM cases "
        "UNION SELECT DISTINCT fecha FROM ot_cases"
    )
    all_dates = sorted({r[0] for r in cur.fetchall() if r[0]})
    conn.close()

    if not all_dates:
        return False, "No cases found in local database."

    processed: list[str] = []
    skipped: list[str] = []
    failed: list[tuple[str, str]] = []

    total = len(all_dates)
    for idx, target_date in enumerate(all_dates, start=1):
        try:
            d = date.fromisoformat(target_date)
        except ValueError:
            failed.append((target_date, "invalid date format"))
            continue

        week_num = d.isocalendar()[1]
        month_str = d.strftime("%Y-%m")
        week_str = f"Week-{week_num:02d}"
        day_dir = os.path.join(productions_dir, month_str, week_str, target_date)
        filename = f"{safe_name}_Production_{target_date}.xlsx"
        out_path = os.path.join(day_dir, filename)

        # Skip if existing file opens successfully (considered valid)
        if os.path.exists(out_path):
            try:
                _check_wb = openpyxl.load_workbook(out_path, read_only=True)
                _check_wb.close()
                skipped.append(target_date)
                if progress_cb:
                    progress_cb(idx, total, f"skip {target_date} (already uploaded)")
                continue
            except Exception:
                log_event("sharepoint_sync",
                          f"bulk: corrupt file at {out_path}, overwriting",
                          level="WARN")

        if progress_cb:
            progress_cb(idx, total, f"building {target_date}")

        try:
            os.makedirs(day_dir, exist_ok=True)
            cases, ot_cases, downtimes, total_cases_pct, total_downtime_min = \
                _get_daily_data(target_date)
            year = int(target_date[:4])
            month = int(target_date[5:7])
            monthly_rows = _get_monthly_data(year, month)
            ue_cases, ue_total, reg_type_counts, ot_type_counts = \
                _get_ue_and_types(target_date)

            wb = openpyxl.Workbook()
            _build_daily_summary(
                wb, target_date, designer,
                cases, ot_cases, downtimes,
                total_cases_pct, total_downtime_min,
                ue_cases=ue_cases, ue_total=ue_total,
            )
            _build_monthly_summary(wb, year, month, designer, monthly_rows)
            _save_atomic(wb, out_path)

            _update_team_summary(
                productions_dir, designer, target_date,
                total_cases_pct, total_downtime_min,
                len(cases), len(ot_cases),
                ue_total, reg_type_counts, ot_type_counts,
                ue_cases=ue_cases,
                rebuild_dashboard=False,  # defer to single rebuild at end
            )
            processed.append(target_date)
        except Exception as exc:
            log_event("sharepoint_sync",
                      f"bulk: failed {target_date}: {exc}",
                      level="ERROR")
            failed.append((target_date, str(exc)))

    # Single Dashboard rebuild at end (stagger + cooldown + verify apply)
    if progress_cb:
        progress_cb(total, total, "rebuilding dashboard")
    try:
        _rebuild_dashboard_file(productions_dir, date.today().isoformat())
    except Exception as exc:
        log_event("sharepoint_sync",
                  f"bulk: dashboard rebuild failed: {exc}",
                  level="WARN")

    parts = [
        f"Processed: {len(processed)} day(s)",
        f"Skipped (already uploaded): {len(skipped)} day(s)",
        f"Failed: {len(failed)} day(s)",
    ]
    if failed:
        preview = "\n".join(f"  {d}: {msg}" for d, msg in failed[:10])
        parts.append("\nFailures:\n" + preview)
        if len(failed) > 10:
            parts.append(f"  ... +{len(failed) - 10} more")
    return (len(failed) == 0), "\n".join(parts)


# ── Audit + historical rebuild ──────────────────────────────────────────────

def _collect_summary_rows(productions_dir: str) -> dict:
    """Return {date: {designer: summary_row_values_tuple}} from all summary files."""
    import glob as _glob
    summary_files = sorted(
        _glob.glob(os.path.join(productions_dir, "*", "_Summary.xlsx"))
    )
    summary_files += sorted(
        _glob.glob(os.path.join(productions_dir, "_Summary_*.xlsx"))
    )

    rows_by_date: dict[str, dict[str, tuple]] = {}
    for sf in summary_files:
        try:
            base = os.path.basename(sf)
            if base == "_Summary.xlsx":
                designer_name = os.path.basename(os.path.dirname(sf)).replace("_", " ")
            else:
                designer_name = base[len("_Summary_"):-len(".xlsx")].replace("_", " ")
            wb = openpyxl.load_workbook(sf, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                rows_by_date.setdefault(str(row[0]), {})[designer_name] = row
            wb.close()
        except Exception as exc:
            log_event("sharepoint_sync",
                      f"audit: could not read summary {sf}: {exc}",
                      level="WARN")
    return rows_by_date


def audit_dashboard_vs_summaries(productions_dir: str | None = None
                                 ) -> tuple[dict, str]:
    """Compare each _Dashboard_YYYY-MM-DD.xlsx snapshot against the per-designer
    summaries for the same date. Reports designers missing from a snapshot or
    whose snapshot row is stale vs the summary.

    Returns (issues_by_date, report_text). issues_by_date is {date: {
        "missing_in_snapshot": [designers],
        "missing_in_summary":  [designers],
        "value_mismatch":      [(designer, field, snapshot_val, summary_val)],
    }}.
    """
    if not _OPENPYXL_OK:
        return {}, f"openpyxl load error: {_OPENPYXL_ERROR}"

    if productions_dir is None:
        cfg = load_config()
        export_folder = (cfg.get("export_folder", "") or "").strip()
        if not export_folder or not os.path.isdir(export_folder):
            return {}, "Export folder not configured or not found."
        productions_dir = os.path.join(export_folder, "Productions")

    snapshots_dir = os.path.join(productions_dir, "Dashboards")
    if not os.path.isdir(snapshots_dir):
        return {}, f"No snapshots folder at {snapshots_dir}"

    import glob as _glob
    snapshot_files = sorted(
        _glob.glob(os.path.join(snapshots_dir, "_Dashboard_*.xlsx"))
    )
    if not snapshot_files:
        return {}, "No dashboard snapshots found."

    rows_by_date = _collect_summary_rows(productions_dir)

    issues: dict[str, dict] = {}
    # Dashboard sheet columns: 1 Designer, 2 Cases%, 3 DT%, 4 Total%,
    # 5 UE(Cases), 6 UE(Total), 7 RegCases, ...
    # Summary row cols (newest schema, 32): 1 Date, 2 Week, 3 Cases%, 4 DT%,
    # 5 Total%, 6 RegCases, 7 OTCases, 8 UE(Total), 9 UE(Cases).
    FIELDS_TO_CHECK = [
        ("Reg Cases",    6,  5),   # (label, dash_idx0, sum_idx0)
        ("Cases %",      1,  2),
        ("UE (Cases)",   4,  8),
        ("UE (Total)",   5,  7),
    ]

    def _norm(v):
        if v is None:
            return None
        if isinstance(v, str):
            s = v.strip()
            if s.endswith("%"):
                try:
                    return round(float(s.rstrip("%")), 1)
                except ValueError:
                    return s
            try:
                return round(float(s), 2)
            except ValueError:
                return s
        try:
            return round(float(v), 2)
        except (TypeError, ValueError):
            return v

    for snap_path in snapshot_files:
        fname = os.path.basename(snap_path)
        # _Dashboard_YYYY-MM-DD.xlsx
        try:
            snap_date = fname[len("_Dashboard_"):-len(".xlsx")]
        except Exception:
            continue

        try:
            wb = openpyxl.load_workbook(snap_path, read_only=True, data_only=True)
        except Exception as exc:
            issues[snap_date] = {"read_error": str(exc)}
            continue

        try:
            ws = wb["Dashboard"] if "Dashboard" in wb.sheetnames else wb.active
        except Exception:
            ws = wb.active

        snap_designers: dict[str, tuple] = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            if name.lower() in ("designer", "team total", "total"):
                continue
            snap_designers[name] = row
        wb.close()

        summary_map = rows_by_date.get(snap_date, {})
        snap_names = set(snap_designers.keys())
        sum_names = set(summary_map.keys())

        missing_in_snap = sorted(sum_names - snap_names)
        missing_in_sum = sorted(snap_names - sum_names)

        mismatches: list[tuple] = []
        for name in sorted(snap_names & sum_names):
            snap_row = snap_designers[name]
            sum_row = summary_map[name]
            for label, d_idx, s_idx in FIELDS_TO_CHECK:
                sv = _norm(snap_row[d_idx]) if d_idx < len(snap_row) else None
                mv = _norm(sum_row[s_idx])  if s_idx < len(sum_row)  else None
                if sv != mv:
                    mismatches.append((name, label, sv, mv))

        if missing_in_snap or missing_in_sum or mismatches:
            issues[snap_date] = {
                "missing_in_snapshot": missing_in_snap,
                "missing_in_summary":  missing_in_sum,
                "value_mismatch":      mismatches,
            }

    # Build report text
    if not issues:
        return issues, (
            f"Audit OK — checked {len(snapshot_files)} snapshot(s). "
            f"No discrepancies found."
        )

    lines = [f"Audit found issues in {len(issues)} of {len(snapshot_files)} snapshot(s):"]
    for d in sorted(issues.keys()):
        info = issues[d]
        lines.append(f"\n[{d}]")
        if "read_error" in info:
            lines.append(f"  ! cannot read snapshot: {info['read_error']}")
            continue
        if info["missing_in_snapshot"]:
            lines.append(f"  Missing in snapshot (present in summary):")
            for n in info["missing_in_snapshot"]:
                lines.append(f"    - {n}")
        if info["missing_in_summary"]:
            lines.append(f"  In snapshot but no summary row:")
            for n in info["missing_in_summary"]:
                lines.append(f"    - {n}")
        if info["value_mismatch"]:
            lines.append(f"  Value mismatches (snapshot → summary):")
            for n, field, sv, mv in info["value_mismatch"][:8]:
                lines.append(f"    - {n} / {field}: {sv} → {mv}")
            if len(info["value_mismatch"]) > 8:
                lines.append(f"    ... +{len(info['value_mismatch']) - 8} more")
    return issues, "\n".join(lines)


def rebuild_historical_dashboards(dates: list[str] | None = None,
                                  productions_dir: str | None = None,
                                  progress_cb=None) -> tuple[bool, str]:
    """Regenerate _Dashboard_{date}.xlsx snapshots for the given dates (or
    every date found in summaries if None). The live _Dashboard.xlsx is
    never touched. Existing snapshots are overwritten.
    """
    if not _OPENPYXL_OK:
        return False, f"openpyxl load error: {_OPENPYXL_ERROR}"

    if productions_dir is None:
        cfg = load_config()
        export_folder = (cfg.get("export_folder", "") or "").strip()
        if not export_folder or not os.path.isdir(export_folder):
            return False, "Export folder not configured or not found."
        productions_dir = os.path.join(export_folder, "Productions")

    if dates is None:
        rows_by_date = _collect_summary_rows(productions_dir)
        dates = sorted(rows_by_date.keys())
    else:
        dates = sorted(set(dates))

    if not dates:
        return False, "No dates to rebuild."

    rebuilt: list[str] = []
    failed: list[tuple[str, str]] = []
    total = len(dates)
    for idx, d_str in enumerate(dates, start=1):
        if progress_cb:
            progress_cb(idx, total, f"rebuilding {d_str}")
        try:
            _rebuild_dashboard_file(
                productions_dir, d_str,
                skip_live=True,
                skip_cache_check=True,
                force_snapshot=True,
            )
            rebuilt.append(d_str)
        except Exception as exc:
            log_event("sharepoint_sync",
                      f"historical rebuild failed {d_str}: {exc}",
                      level="WARN")
            failed.append((d_str, str(exc)))

    parts = [
        f"Rebuilt: {len(rebuilt)} snapshot(s)",
        f"Failed: {len(failed)} snapshot(s)",
    ]
    if failed:
        preview = "\n".join(f"  {d}: {msg}" for d, msg in failed[:10])
        parts.append("\nFailures:\n" + preview)
        if len(failed) > 10:
            parts.append(f"  ... +{len(failed) - 10} more")
    return (len(failed) == 0), "\n".join(parts)
