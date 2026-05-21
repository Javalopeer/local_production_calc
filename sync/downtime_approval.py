# -*- coding: utf-8 -*-
"""
Downtime export — per-designer files + consolidated history.

Architecture (scales to 50+ designers without conflicts):
  1. Designer adds a downtime → status = 'pending' in local DB.
  2. The user pastes the DT into Teams using the Copy button. The supervisor
     reacts with a like/heart/thumbs-down in Teams. Then the user, back in
     their own app, marks the DT Approved or Rejected manually.
  3. App writes ONLY this designer's rows to their own file:
       Downtime/designers_dt/_DT_<DesignerName>.xlsx
     No other designer ever touches this file → zero write conflicts.
  4. App reads ALL designers_dt/_DT_*.xlsx files and rebuilds a read-only
     consolidated history view at:
       Downtime/_Downtime_Approvals.xlsx
     Used by leads to audit who logged what and the current status.
"""

import glob
import os
import time
from datetime import datetime

from db.database import get_connection
from sync.app_config import load_config, get_excel_sheet_password
from sync.app_logger import log_event

STATUS_PENDING  = "pending"
STATUS_APPROVED = "approved"
STATUS_REJECTED = "rejected"

_DT_WRITE_JITTER   = (1.0, 4.0)  # random backoff (s) on individual designer file write retry
_DT_CONSOL_JITTER  = (0.5, 2.0)  # random backoff (s) on consolidated file write retry

CONSOLIDATED_FILE = "_Downtime_Approvals.xlsx"

# Cache: track the latest mtime of _DT_*.xlsx files to skip unnecessary rebuilds
_last_rebuild_max_mtime: float = 0.0

# Track failed exports so they can be retried on next poll cycle
_pending_retry_designer: str = ""

_OPENPYXL_OK = False
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table as XlTable, TableStyleInfo
    _OPENPYXL_OK = True
except ImportError:
    log_event("downtime_approval", "openpyxl not installed; approval excel features disabled", level="WARN")

_SHEET_PASSWORD = get_excel_sheet_password()


# ── helpers ──────────────────────────────────────────────────────────────────

def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


_README_TEXT = (
    "Downtime folder\n"
    "================\n\n"
    "For team-wide history, open:\n"
    "  _Downtime_Approvals.xlsx     — every DT, every designer, every status.\n\n"
    "Do NOT edit:\n"
    "  _Downtime_Approvals.xlsx     — rebuilt automatically from designer files.\n"
    "  designers_dt\\_DT_<name>.xlsx  — each designer's own xlsx.\n\n"
    "Manual edits will be overwritten by the next rebuild.\n"
    "Approvals/rejections are now made directly inside each user's app after\n"
    "they paste the DT into the supervisor's Teams chat.\n"
)


def _ensure_downtime_readme(downtime_dir: str) -> None:
    """Write README.txt once so leads know which file is for them."""
    readme_path = os.path.join(downtime_dir, "README.txt")
    if os.path.exists(readme_path):
        return
    try:
        with open(readme_path, "w", encoding="utf-8") as fh:
            fh.write(_README_TEXT)
    except OSError as exc:
        log_event("downtime_approval", f"could not write README.txt: {exc}", level="WARN")


def _get_downtime_dir() -> str | None:
    """Return the Downtime/ folder path, creating it if needed."""
    cfg = load_config()
    folder = cfg.get("export_folder", "").strip()
    if not folder:
        return None
    if not os.path.isdir(folder):
        return None
    downtime_dir = os.path.join(folder, "Downtime")
    try:
        os.makedirs(downtime_dir, exist_ok=True)
    except OSError:
        return None
    _ensure_downtime_readme(downtime_dir)
    return downtime_dir


def get_approval_path() -> str | None:
    """Return the full path to the consolidated approval file."""
    d = _get_downtime_dir()
    return os.path.join(d, CONSOLIDATED_FILE) if d else None


def _designer_file_path(designer_name: str) -> str | None:
    """Return the path for this designer's individual file inside designers_dt/."""
    d = _get_downtime_dir()
    if not d:
        return None
    designers_dir = os.path.join(d, "designers_dt")
    try:
        os.makedirs(designers_dir, exist_ok=True)
    except OSError:
        return None
    safe_name = "".join(c for c in designer_name.strip() if c.isalnum() or c in " _-").strip()
    if not safe_name:
        safe_name = "Unknown"
    return os.path.join(designers_dir, f"_DT_{safe_name}.xlsx")


def _save_workbook(wb, path: str, retries: int = 6) -> bool:
    """Save a workbook with retries + random jitter for OneDrive locks."""
    import random
    for attempt in range(retries):
        try:
            wb.save(path)
            return True
        except PermissionError:
            if attempt < retries - 1:
                time.sleep(random.uniform(*_DT_WRITE_JITTER))
        except Exception as exc:
            log_event("downtime_approval", f"save failed: {exc}", level="WARN")
            return False
    log_event("downtime_approval", f"file locked after {retries} attempts: {os.path.basename(path)}", level="WARN")
    return False


def _read_workbook(path: str, retries: int = 4):
    """Read a workbook with retries. Returns workbook or None."""
    import random
    for attempt in range(retries):
        try:
            return openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            if attempt == retries - 1:
                log_event("downtime_approval", f"read workbook failed for {os.path.basename(path)} after {retries} tries: {exc}", level="WARN")
            if attempt < retries - 1:
                time.sleep(random.uniform(*_DT_CONSOL_JITTER))
    return None


# ── export (per-designer file) ───────────────────────────────────────────────

def export_pending_downtimes(designer_name: str) -> bool:
    """Write this designer's downtimes to their own _DT_<name>.xlsx file,
    then rebuild the consolidated file.

    Returns True on success.  On failure, sets _pending_retry_designer so
    the next poll cycle will retry automatically.
    """
    global _pending_retry_designer
    if not _OPENPYXL_OK:
        return False

    designer_path = _designer_file_path(designer_name)
    if not designer_path:
        return False

    # ── Fetch data from DB ───────────────────────────────────────────────────
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, fecha, hora_inicio, hora_fin, duracion, razon, status,
               responded_by, responded_at
        FROM downtimes
        ORDER BY fecha DESC, hora_inicio DESC
    """)
    all_rows = cur.fetchall()
    # IDs we'll mark as synced_to_excel after a successful save
    all_ids = [r[0] for r in all_rows]
    conn.close()

    pending = [r for r in all_rows if r[6] == STATUS_PENDING]

    # ── Build designer workbook ──────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # -- Pending sheet --
    ws = wb.active
    ws.title = "Pending"
    headers = ["Designer", "ID", "Date", "Start", "End", "Duration (min)", "Reason", "Status",
               "Responded By", "Responded At"]
    _write_header(ws, headers, PatternFill("solid", fgColor="2D89EF"))

    pend_fill = PatternFill("solid", fgColor="FFF9C4")
    row_idx = 2
    for db_row in pending:
        _write_data_row(ws, row_idx, designer_name, db_row, pend_fill)
        row_idx += 1
    if row_idx == 2:
        ws.cell(2, 1, "(no pending requests)")
        ws.merge_cells("A2:J2")
        ws.cell(2, 1).alignment = Alignment(horizontal="center")

    _set_column_widths(ws)

    # -- History sheet --
    ws_hist = wb.create_sheet("History")
    _write_header(ws_hist, headers, PatternFill("solid", fgColor="4CAF50"))

    status_fills = {
        STATUS_APPROVED: PatternFill("solid", fgColor="E8F5E9"),
        STATUS_REJECTED: PatternFill("solid", fgColor="FFEBEE"),
        STATUS_PENDING:  PatternFill("solid", fgColor="FFF9C4"),
    }
    hist_row = 2
    for db_row in all_rows:
        row_status = str(db_row[6] or STATUS_PENDING).lower()
        fill = status_fills.get(row_status, PatternFill("solid", fgColor="F8F8F8"))
        _write_data_row(ws_hist, hist_row, designer_name, db_row, fill)
        hist_row += 1
    if hist_row == 2:
        ws_hist.cell(2, 1, "(no downtime history)")
        ws_hist.merge_cells("A2:J2")
        ws_hist.cell(2, 1).alignment = Alignment(horizontal="center")

    _set_column_widths(ws_hist)

    # ── Save designer file (only this designer writes here → no conflicts) ──
    if not _save_workbook(wb, designer_path):
        _pending_retry_designer = designer_name
        print(f"[downtime_approval] Export failed, will retry on next poll cycle.")
        return False
    print(f"[downtime_approval] Designer file saved: {os.path.basename(designer_path)}")

    # ── Rebuild consolidated file ────────────────────────────────────────────
    _rebuild_consolidated(force=True)

    # ── Mark every exported row as synced_to_excel ──────────────────────────
    # The designer's xlsx is the source of truth other machines read from, so
    # once the save returned True every row in `all_rows` has reached the
    # shared folder. The retry worker uses this flag to know it can stop
    # retrying these IDs.
    if all_ids:
        try:
            conn = get_connection()
            placeholders = ",".join("?" for _ in all_ids)
            conn.execute(
                f"UPDATE downtimes SET synced_to_excel = 1, last_sync_error = '' "
                f"WHERE id IN ({placeholders})",
                all_ids,
            )
            conn.commit()
            conn.close()
        except Exception as exc:
            print(f"[downtime_approval] Could not flag rows synced_to_excel: {exc}")

    _pending_retry_designer = ""
    return True


# ── consolidated file rebuild ────────────────────────────────────────────────

def _rebuild_consolidated(force: bool = False) -> bool:
    """Read all _DT_*.xlsx files and build _Downtime_Approvals.xlsx.

    Skips the rebuild if no _DT_*.xlsx file has changed since the last
    rebuild (based on file mtimes), unless force=True.
    """
    global _last_rebuild_max_mtime

    downtime_dir = _get_downtime_dir()
    if not downtime_dir:
        return False

    consolidated_path = os.path.join(downtime_dir, CONSOLIDATED_FILE)

    # ── Check if any designer file changed since last rebuild ────────────────
    designers_dir = os.path.join(downtime_dir, "designers_dt")
    os.makedirs(designers_dir, exist_ok=True)
    pattern = os.path.join(designers_dir, "_DT_*.xlsx")
    dt_files = [f for f in glob.glob(pattern) if os.path.basename(f) != CONSOLIDATED_FILE]

    if not force and dt_files:
        max_mtime = max(os.path.getmtime(f) for f in dt_files)
        if max_mtime <= _last_rebuild_max_mtime and os.path.exists(consolidated_path):
            print("[downtime_approval] No designer files changed, skipping rebuild.")
            return True

    # ── Collect rows from each designer file ────────────────────────────────
    # Each designer's _DT_<name>.xlsx is the source of truth for their own
    # rows; the consolidated file is just a read-only view that aggregates
    # them. Supervisor edits used to live in the consolidated file under the
    # old Power Automate flow — that workflow was retired, so we no longer
    # try to preserve them here.
    all_pending = []
    all_history = []

    def _pad10(r):
        tup = list(r[:10])
        while len(tup) < 10:
            tup.append("")
        return tuple(tup)

    for dt_file in dt_files:
        wb = _read_workbook(dt_file, retries=2)
        if wb is None:
            continue
        try:
            ws = wb.active
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r or r[0] is None:
                    continue
                row_designer = str(r[0] or "").strip()
                if row_designer.startswith("("):
                    continue
                all_pending.append(_pad10(r))

            if "History" in wb.sheetnames:
                ws_hist = wb["History"]
                for r in ws_hist.iter_rows(min_row=2, values_only=True):
                    if not r or r[0] is None:
                        continue
                    row_designer = str(r[0] or "").strip()
                    if row_designer.startswith("("):
                        continue
                    all_history.append(_pad10(r))

            wb.close()
        except Exception as exc:
            print(f"[downtime_approval] Error reading {os.path.basename(dt_file)}: {exc}")

    # ── Build consolidated workbook ──────────────────────────────────────────
    wb = openpyxl.Workbook()

    # -- Pending Approvals sheet --
    ws = wb.active
    ws.title = "Downtime Approvals"
    ws.sheet_view.showGridLines = False
    headers = ["Designer", "ID", "Date", "Start", "End", "Duration (min)", "Reason", "Status",
               "Responded By", "Responded At"]
    _write_header(ws, headers, PatternFill("solid", fgColor="2D89EF"))

    note = ws.cell(1, 12, "► Read-only history. Each designer marks their own DTs in their app.")
    note.font = Font(italic=True, color="555555")

    _locked   = Protection(locked=True)
    _unlocked = Protection(locked=False)

    pend_fill  = PatternFill("solid", fgColor="FFF9C4")
    other_fill = PatternFill("solid", fgColor="F8F8F8")

    row_idx = 2
    for r in all_pending:
        for c_idx, val in enumerate(r, 1):
            cell = ws.cell(row_idx, c_idx, val)
            cell.alignment = Alignment(
                horizontal="left" if c_idx in (1, 7) else "center",
                vertical="center"
            )
            cell.fill   = pend_fill
            cell.border = _thin()
            cell.protection = _unlocked if c_idx == 8 else _locked
        row_idx += 1

    if row_idx == 2:
        # Add a placeholder row so the Excel Table always exists
        placeholder = ["—", 0, "—", "—", "—", 0, "(no pending requests)", "—", "", ""]
        for c_idx, val in enumerate(placeholder, 1):
            cell = ws.cell(2, c_idx, val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill   = other_fill
            cell.border = _thin()
        row_idx = 3

    # Dropdown validation on Status column
    dv = DataValidation(
        type="list",
        formula1='"pending,approved,rejected"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error="Please select: pending, approved, or rejected.",
    )
    ws.add_data_validation(dv)
    dv.sqref = f"H2:H{row_idx - 1}"

    _set_column_widths(ws)
    ws.freeze_panes = "A2"

    # Column filter dropdowns so leads can filter by Date / Reason / Status /
    # Designer before bulk-approving a time window.
    ws.auto_filter.ref = f"A1:J{row_idx - 1}"

    # Excel Table so leads can filter/sort downtimes for review. Ref includes
    # Responded By / Responded At columns so they're inside the table area.
    tab = XlTable(
        displayName="DowntimeApprovals",
        ref=f"A1:J{row_idx - 1}",
    )
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True,
    )
    ws.add_table(tab)

    # -- History sheet --
    ws_hist = wb.create_sheet("History")
    ws_hist.sheet_view.showGridLines = False
    _write_header(ws_hist, headers, PatternFill("solid", fgColor="4CAF50"))

    status_fills = {
        STATUS_APPROVED: PatternFill("solid", fgColor="E8F5E9"),
        STATUS_REJECTED: PatternFill("solid", fgColor="FFEBEE"),
        STATUS_PENDING:  PatternFill("solid", fgColor="FFF9C4"),
    }
    hist_row = 2
    for r in all_history:
        row_status = str(r[7] if len(r) > 7 else STATUS_PENDING).lower()
        fill = status_fills.get(row_status, other_fill)
        for c_idx, val in enumerate(r, 1):
            cell = ws_hist.cell(hist_row, c_idx, val)
            cell.alignment = Alignment(
                horizontal="left" if c_idx in (1, 7) else "center",
                vertical="center"
            )
            cell.fill   = fill
            cell.border = _thin()
            cell.protection = _locked
        hist_row += 1

    if hist_row == 2:
        ws_hist.cell(2, 1, "(no downtime history)")
        ws_hist.merge_cells("A2:J2")
        ws_hist.cell(2, 1).alignment = Alignment(horizontal="center")

    ws_hist.protection.sheet    = True
    ws_hist.protection.password = _SHEET_PASSWORD
    ws_hist.protection.autoFilter = False  # allow leads to use the dropdowns
    ws_hist.protection.sort       = False
    ws_hist.protection.enable()
    _set_column_widths(ws_hist)
    ws_hist.freeze_panes = "A2"
    if hist_row > 2:
        ws_hist.auto_filter.ref = f"A1:J{hist_row - 1}"

    # ── Save consolidated — if it fails, retry on next cycle ────────────────
    if _save_workbook(wb, consolidated_path):
        # Update mtime cache so next call skips rebuild if nothing changed
        if dt_files:
            _last_rebuild_max_mtime = max(os.path.getmtime(f) for f in dt_files if os.path.exists(f))
        print(f"[downtime_approval] Consolidated file rebuilt with {row_idx - 2} pending row(s).")
        return True
    else:
        print("[downtime_approval] Could not update consolidated file (a lead may have it open). Will retry next cycle.")
        return False


# ── shared formatting helpers ────────────────────────────────────────────────

def _write_header(ws, headers: list, fill: PatternFill):
    hdr_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill      = fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin()


def _write_data_row(ws, row_idx: int, designer: str, db_row, fill: PatternFill):
    # db_row may have 7 (legacy) or 9 (with responded_by/at) fields;
    # pad to 9 so the sheet always has the two extra columns.
    row_vals = list(db_row)
    while len(row_vals) < 9:
        row_vals.append("")
    vals = [designer] + row_vals[:9]
    for c_idx, val in enumerate(vals, 1):
        cell = ws.cell(row_idx, c_idx, val)
        cell.alignment = Alignment(
            horizontal="left" if c_idx in (1, 7) else "center",
            vertical="center"
        )
        cell.fill   = fill
        cell.border = _thin()


def _set_column_widths(ws):
    widths = {"A": 22, "B": 8, "C": 14, "D": 10, "E": 10, "F": 16, "G": 38, "H": 14,
              "I": 22, "J": 20}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 22


