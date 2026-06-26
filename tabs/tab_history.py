from PySide6.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout,
    QPushButton, QLineEdit, QTableWidget, QTableWidgetItem,
    QDateEdit, QComboBox, QFileDialog, QMessageBox, QCheckBox,
    QApplication, QHeaderView,
)
from PySide6.QtCore import QDate, Qt
from PySide6.QtGui import QColor, QFont, QBrush
from db.database import get_connection
from .utils import (
    load_units_eq_data, calculate_equivalent_units,
    DAILY_BASE_MINUTES,
)
from .theme_table_utils import (
    apply_table_theme, CLR_FG_LIGHT, CLR_FG_DARK,
    get_light_theme_colors, light_row_bg, light_header_bg, light_header_fg, mix_hex,
)
from .theme_palette import apply_fluent_modal_palette
import csv

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class HistoryTab(QWidget):
    def __init__(self):
        super().__init__()
        self.all_cases = []
        self.filtered_cases = []
        self.units_eq = load_units_eq_data()
        self.current_page = 1
        self.items_per_page = 50
        self.total_pages = 1
        self.init_ui()
        self.load_all_cases()

    def init_ui(self):
        from PySide6.QtWidgets import QFrame
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(16, 14, 16, 10)
        main_layout.setSpacing(12)

        # Title — Production-style header.
        title = QLabel("Case History")
        title.setStyleSheet(
            "color: #E6EDF3; font-size: 18px; font-weight: 800;"
            " letter-spacing: 0.3px;"
        )
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title)

        # Stats KPI cards — mirrors Production tab. Legacy stats_label /
        # stats_label2 are kept hidden so the existing setText path stays
        # alive; new card-level values come from _kpi_* labels.
        from PySide6.QtWidgets import QFrame as _QF_h
        try:
            from .tabler_icons import TablerIcon as _TI_kpi
            from PySide6.QtGui import QColor as _QC_kpi
        except Exception:
            _TI_kpi = None

        def _kpi(label_text: str, value_widget: QLabel, icon_svg: str):
            card = _QF_h()
            card.setObjectName("kpiCard")
            card.setStyleSheet(
                "#kpiCard { background: #0D1117; border: 1px solid #21262D;"
                " border-radius: 10px; }"
                "QLabel { background: transparent; border: none; }"
            )
            h = QHBoxLayout(card)
            h.setContentsMargins(14, 10, 14, 10)
            h.setSpacing(10)
            if _TI_kpi is not None and icon_svg:
                ic_lbl = QLabel()
                ic_lbl.setFixedSize(34, 34)
                ic_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
                def _apply_hist_kpi_icon(is_light: bool, _l=ic_lbl, _svg=icon_svg):
                    try:
                        from .theme_palette import palette
                        col = palette(is_light)["text"]
                    except Exception:
                        col = "#1F2328" if is_light else "#FFFFFF"
                    _l.setPixmap(
                        _TI_kpi(_svg).icon(color=_QC_kpi(col)).pixmap(20, 20)
                    )
                ic_lbl.apply_palette = _apply_hist_kpi_icon
                try:
                    from qfluentwidgets.common.style_sheet import isDarkTheme
                    _apply_hist_kpi_icon(not isDarkTheme())
                except Exception:
                    _apply_hist_kpi_icon(False)
                h.addWidget(ic_lbl, 0, Qt.AlignmentFlag.AlignVCenter)
            col = QVBoxLayout(); col.setSpacing(0)
            l = QLabel(label_text)
            l.setAlignment(Qt.AlignmentFlag.AlignCenter)
            l.setStyleSheet("color: #8B949E; font-size: 11px; font-weight: 600;")
            value_widget.setAlignment(Qt.AlignmentFlag.AlignCenter)
            value_widget.setStyleSheet(
                "color: #E6EDF3; font-size: 18px; font-weight: 500;"
            )
            col.addWidget(l); col.addWidget(value_widget)
            wrap = QVBoxLayout()
            wrap.addStretch(1); wrap.addLayout(col); wrap.addStretch(1)
            outer = QHBoxLayout()
            outer.addStretch(1); outer.addLayout(wrap); outer.addStretch(1)
            h.addLayout(outer, 1)
            return card

        self._kpi_cases = QLabel("—")
        self._kpi_time = QLabel("—")
        self._kpi_value = QLabel("—")
        self._kpi_ue = QLabel("—")

        kpi_row = QHBoxLayout(); kpi_row.setSpacing(10)
        kpi_row.addWidget(_kpi("Cases", self._kpi_cases,
                                "tabler_clipboard_text.svg"), 1)
        kpi_row.addWidget(_kpi("Time", self._kpi_time,
                                "tabler_clock.svg"), 1)
        kpi_row.addWidget(_kpi("Value", self._kpi_value,
                                "tabler_percentage_30.svg"), 1)
        kpi_row.addWidget(_kpi("Equivalent Units", self._kpi_ue,
                                "tabler_congruent_to.svg"), 1)
        main_layout.addLayout(kpi_row)

        # Legacy labels still referenced by filter_data — kept hidden.
        self.stats_label = QLabel("")
        self.stats_label.setVisible(False)
        self.stats_label2 = QLabel("")
        self.stats_label2.setVisible(False)

        # ── Filter card — single row mirroring Production tab ─────────
        try:
            from .widgets import _icon_url as _icu_h
            _chev_h = _icu_h("tabler_chevron_down.svg")
        except Exception:
            _chev_h = ""
        _input_css_h = (
            "QLineEdit, QDateEdit, QComboBox { background: #161B22;"
            " border: 1px solid #30363D; border-radius: 6px;"
            " padding: 4px 8px; color: #E6EDF3; font-size: 11px;"
            " min-height: 26px; }"
            "QComboBox::drop-down, QDateEdit::drop-down {"
            " subcontrol-origin: padding; subcontrol-position: right center;"
            " width: 22px; border: none; }"
            f"QComboBox::down-arrow, QDateEdit::down-arrow {{"
            f" image: url({_chev_h}); width: 12px; height: 12px; }}"
        )

        def _col_label(text):
            l = QLabel(text)
            l.setStyleSheet(
                "color: #C9D1D9; font-size: 12px; font-weight: 700;"
                " background: transparent;"
            )
            return l

        filter_card = QFrame()
        filter_card.setObjectName("filterCard")
        filter_card.setStyleSheet(
            "#filterCard { background: #0D1117; border: 1px solid #21262D;"
            " border-radius: 10px; }"
            "QLabel { background: transparent; border: none;"
            "  color: #C9D1D9; font-size: 11px; font-weight: 600; }"
            "QCheckBox { color: #C9D1D9; font-size: 11px; spacing: 6px;"
            "  background: transparent; }"
        )
        fb = QVBoxLayout(filter_card)
        fb.setContentsMargins(14, 12, 14, 18)
        fb.setSpacing(10)

        try:
            from .widgets import DateEditWithShortcut as _DateEditH
        except Exception:
            _DateEditH = QDateEdit
        try:
            from .tabler_icons import TablerIcon as _TI_h
            from PySide6.QtGui import QAction as _QA_h, QColor as _QC_h
            from PySide6.QtCore import QSize as _QS_h
            _TI_avail = True
        except Exception:
            _TI_avail = False

        def _attach_cal(de):
            if not _TI_avail:
                return
            try:
                le = de.lineEdit() if hasattr(de, "lineEdit") else None
                if le is None:
                    return
                act = _QA_h(
                    _TI_h("tabler_calendar.svg").icon(color=_QC_h("#8B949E")),
                    "", le,
                )
                le.addAction(act, QLineEdit.ActionPosition.LeadingPosition)
            except Exception:
                pass

        def _stacked(label_text, widget, stretch=0):
            """Vertical column: label centered above the input."""
            col = QVBoxLayout(); col.setSpacing(2)
            l = QLabel(label_text)
            l.setAlignment(Qt.AlignmentFlag.AlignCenter)
            l.setStyleSheet(
                "color: #8B949E; font-size: 10px; font-weight: 700;"
                " letter-spacing: 0.5px; background: transparent;"
            )
            col.addWidget(l)
            col.addWidget(widget)
            return col, stretch

        # ── Row 1: Search · Status · From · To · Specific date ────────
        row1 = QHBoxLayout(); row1.setSpacing(10)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search case ID or doctor…")
        self.search_input.setStyleSheet(_input_css_h)
        self.search_input.setFixedHeight(30)
        self.search_input.setMinimumWidth(260)
        self.search_input.textChanged.connect(self.on_filter_changed)
        if _TI_avail:
            try:
                self.search_input.addAction(
                    _QA_h(_TI_h("tabler_search.svg").icon(color=_QC_h("#8B949E")),
                          "", self.search_input),
                    QLineEdit.ActionPosition.LeadingPosition,
                )
            except Exception:
                pass
        col, st = _stacked("Search", self.search_input, 2)
        row1.addLayout(col, st)

        self.status_filter = QComboBox()
        self.status_filter.addItems(["All", "OK", "LOW"])
        self.status_filter.setStyleSheet(_input_css_h)
        self.status_filter.setFixedHeight(30)
        self.status_filter.setMinimumWidth(110)
        self.status_filter.currentTextChanged.connect(self.on_filter_changed)
        col, st = _stacked("Status", self.status_filter, 0)
        row1.addLayout(col, st)

        # From/To removed — only Specific date filters by exact day. The
        # hidden date_from / date_to widgets stay for backend compat
        # (set to a wide range so they don't restrict the result set).
        self.date_from = _DateEditH()
        self.date_from.setDate(QDate(2020, 1, 1))
        self.date_from.hide()
        self.date_to = _DateEditH()
        self.date_to.setDate(QDate(2099, 12, 31))
        self.date_to.hide()

        # Specific date: checkbox + picker side-by-side.
        self.specific_date_check = QCheckBox("Specific date")
        self.specific_date_check.stateChanged.connect(self.on_specific_date_toggled)
        try:
            from .widgets import _icon_url as _icu_ck
            _check_url = _icu_ck("tabler_check.svg")
        except Exception:
            _check_url = ""
        self.specific_date_check.setStyleSheet(
            "QCheckBox { color: #C9D1D9; font-size: 11px; spacing: 8px;"
            "  background: transparent; }"
            "QCheckBox::indicator { width: 16px; height: 16px;"
            "  background: #161B22; border: 1px solid #30363D;"
            "  border-radius: 4px; }"
            "QCheckBox::indicator:hover { border-color: #58606A; }"
            f"QCheckBox::indicator:checked {{"
            f"  background: #1e63e4; border-color: #1e63e4;"
            f"  image: url({_check_url}); }}"
        )
        self.specific_date = _DateEditH()
        self.specific_date.setDate(QDate.currentDate())
        self.specific_date.setCalendarPopup(True)
        self.specific_date.setDisplayFormat("yyyy-MM-dd")
        self.specific_date.setFixedHeight(30)
        self.specific_date.setMinimumWidth(150)
        self.specific_date.setStyleSheet(_input_css_h)
        # Always editable so users can change the date directly. The
        # checkbox controls whether the date filter is APPLIED — not
        # whether the input is interactable.
        self.specific_date.setEnabled(True)
        self.specific_date.dateChanged.connect(self.on_filter_changed)
        _attach_cal(self.specific_date)
        sd_inner = QWidget()
        sd_inner_lay = QHBoxLayout(sd_inner)
        sd_inner_lay.setContentsMargins(0, 0, 0, 0)
        sd_inner_lay.setSpacing(8)
        sd_inner_lay.addWidget(self.specific_date_check)
        sd_inner_lay.addWidget(self.specific_date)
        col, st = _stacked("Date", sd_inner, 0)
        row1.addLayout(col, st)

        fb.addLayout(row1)

        # Region/Type/Count moved into a popup opened by the Filters
        # button. Combos still exist (hidden) — they're read directly by
        # the filter pipeline so we just write into them on popup Apply.
        self.filter_region = QComboBox(); self.filter_region.hide()
        self.filter_region.currentTextChanged.connect(self.on_filter_changed)
        self.filter_type = QComboBox(); self.filter_type.hide()
        self.filter_type.currentTextChanged.connect(self.on_filter_changed)
        self.count_filter = QComboBox()
        self.count_filter.addItems(["All", "Counted", "NC"])
        self.count_filter.hide()
        self.count_filter.currentTextChanged.connect(self.on_filter_changed)

        # ── Row 2: Doctor (left) · Filters · Export CSV ───────────────
        row2 = QHBoxLayout(); row2.setSpacing(10)

        self.filter_doctor = QLineEdit()
        self.filter_doctor.setPlaceholderText("Search doctor…")
        self.filter_doctor.setStyleSheet(_input_css_h)
        self.filter_doctor.setFixedHeight(30)
        self.filter_doctor.setMinimumWidth(180)
        self.filter_doctor.textChanged.connect(self.on_filter_changed)
        col, st = _stacked("Doctor", self.filter_doctor, 1)
        row2.addLayout(col, st)

        # ── Comments toggle chips ──
        self._filter_comments = ""

        def _make_chip(text, icon_svg=None):
            b = QPushButton("  " + text)
            b.setCheckable(True)
            b.setCursor(Qt.PointingHandCursor)
            b.setFixedHeight(30)
            if _TI_avail and icon_svg:
                try:
                    b.setIcon(_TI_h(icon_svg).icon(color=_QC_h("#8B949E")))
                    b.setIconSize(_QS_h(14, 14))
                except Exception:
                    pass
            b.setStyleSheet(
                "QPushButton { background: #161B22; border: 1px solid #30363D;"
                "  color: #C9D1D9; border-radius: 6px; padding: 0 12px;"
                "  font-size: 11px; font-weight: 600; }"
                "QPushButton:hover { border-color: #58606A; }"
                "QPushButton:checked { background: rgba(56,139,253,0.14);"
                "  border-color: #388BFD; color: #58A6FF; }"
            )
            return b

        self._chip_with = _make_chip("With", "tabler_message_circle.svg")
        self._chip_without = _make_chip("Without", "tabler_message_off.svg")
        cmts_inner = QWidget()
        cmts_inner_lay = QHBoxLayout(cmts_inner)
        cmts_inner_lay.setContentsMargins(0, 0, 0, 0)
        cmts_inner_lay.setSpacing(6)
        cmts_inner_lay.addWidget(self._chip_with)
        cmts_inner_lay.addWidget(self._chip_without)
        col, st = _stacked("Comments", cmts_inner, 0)
        row2.addLayout(col, st)

        def _on_cmts_chip(key):
            if key == "with":
                if self._chip_with.isChecked():
                    self._chip_without.setChecked(False)
                    self._filter_comments = "with"
                else:
                    self._filter_comments = ""
            else:
                if self._chip_without.isChecked():
                    self._chip_with.setChecked(False)
                    self._filter_comments = "without"
                else:
                    self._filter_comments = ""
            self.on_filter_changed()
        self._chip_with.clicked.connect(lambda: _on_cmts_chip("with"))
        self._chip_without.clicked.connect(lambda: _on_cmts_chip("without"))

        # ── Counted toggle chips ──
        self._filter_counted = ""
        self._chip_counted = _make_chip("Counted", "tabler_check.svg")
        self._chip_nc = _make_chip("NC", "tabler_x.svg")
        cnt_inner = QWidget()
        cnt_inner_lay = QHBoxLayout(cnt_inner)
        cnt_inner_lay.setContentsMargins(0, 0, 0, 0)
        cnt_inner_lay.setSpacing(6)
        cnt_inner_lay.addWidget(self._chip_counted)
        cnt_inner_lay.addWidget(self._chip_nc)
        col, st = _stacked("Count", cnt_inner, 0)
        row2.addLayout(col, st)

        def _on_cnt_chip(key):
            if key == "counted":
                if self._chip_counted.isChecked():
                    self._chip_nc.setChecked(False)
                    self._filter_counted = "counted"
                else:
                    self._filter_counted = ""
            else:
                if self._chip_nc.isChecked():
                    self._chip_counted.setChecked(False)
                    self._filter_counted = "nc"
                else:
                    self._filter_counted = ""
            self.on_filter_changed()
        self._chip_counted.clicked.connect(lambda: _on_cnt_chip("counted"))
        self._chip_nc.clicked.connect(lambda: _on_cnt_chip("nc"))

        # Filters button — opens popup with Region/Type/Count.
        self._filters_btn = QPushButton("  Filters")
        self._filters_btn.setCursor(Qt.PointingHandCursor)
        self._filters_btn.setFixedHeight(30)
        self._filters_btn.setMinimumWidth(110)
        if _TI_avail:
            try:
                self._filters_btn.setIcon(
                    _TI_h("tabler_filter.svg").icon(color=_QC_h("#58A6FF"))
                )
                self._filters_btn.setIconSize(_QS_h(14, 14))
            except Exception:
                pass
        self._filters_btn.setStyleSheet(
            "QPushButton { background: transparent; border: 1px solid #1e63e4;"
            "  color: #58A6FF; border-radius: 6px; padding: 4px 14px;"
            "  font-weight: 700; font-size: 11px; }"
            "QPushButton:hover { background: rgba(30,99,228,0.10); }"
        )
        self._filters_btn.clicked.connect(self._open_filter_popup)
        filt_col = QVBoxLayout(); filt_col.setSpacing(2)
        _spacer_lblf = QLabel(" ")
        _spacer_lblf.setStyleSheet("font-size: 10px; background: transparent;")
        filt_col.addWidget(_spacer_lblf)
        filt_col.addWidget(self._filters_btn)
        row2.addLayout(filt_col, 0)

        self.export_btn = QPushButton("  Export CSV")
        self.export_btn.setCursor(Qt.PointingHandCursor)
        self.export_btn.setFixedHeight(30)
        self.export_btn.setMinimumWidth(120)
        if _TI_avail:
            try:
                self.export_btn.setIcon(
                    _TI_h("tabler_download.svg").icon(color=_QC_h("#58A6FF"))
                )
                self.export_btn.setIconSize(_QS_h(14, 14))
            except Exception:
                pass
        self.export_btn.setStyleSheet(
            "QPushButton { background: transparent; border: 1px solid #1e63e4;"
            "  color: #58A6FF; border-radius: 6px; padding: 4px 14px;"
            "  font-weight: 700; font-size: 11px; }"
            "QPushButton:hover { background: rgba(30,99,228,0.10); }"
        )
        self.export_btn.clicked.connect(self.export_csv)
        # Stacked with empty label spacer so the button aligns with the
        # input row instead of riding above it.
        export_col = QVBoxLayout(); export_col.setSpacing(2)
        _spacer_lbl = QLabel(" ")
        _spacer_lbl.setStyleSheet("font-size: 10px; background: transparent;")
        export_col.addWidget(_spacer_lbl)
        export_col.addWidget(self.export_btn)
        row2.addLayout(export_col, 0)

        fb.addLayout(row2)
        main_layout.addWidget(filter_card)

        # Legacy attrs kept so resizeEvent / on_specific_date_toggled
        # / responsive logic referenced elsewhere don't blow up.
        self.specific_date_label = QLabel(""); self.specific_date_label.setVisible(False)
        self.filter_row1 = QHBoxLayout()
        self.filter_row2 = QHBoxLayout()
        self.filter_row3 = QHBoxLayout()
        self.specific_row = QHBoxLayout()
        self.specific_row_widget = QWidget()
        self.specific_row_widget.setLayout(self.specific_row)
        self.specific_row_widget.setVisible(False)
        self.filter_row3_widget = QWidget()
        self.filter_row3_widget.setLayout(self.filter_row3)
        self.filter_row3_widget.setVisible(False)
        self.lbl_region = QLabel("Region")
        self.lbl_type = QLabel("Type")
        self.lbl_count = QLabel("Count")
        self.lbl_doctor = QLabel("Doctor")
        for _l in (self.lbl_region, self.lbl_type, self.lbl_count, self.lbl_doctor):
            _l.setVisible(False)
        self.specific_in_row1 = True
        self.row2_compact = False

        # Table
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(False)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(32)
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "CASE ID", "DOCTOR", "REGION", "TYPE", "DATE",
            "TIME", "STD", "EFF %", "VALUE %",
        ])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setShowGrid(True)
        self.table.setGridStyle(Qt.PenStyle.SolidLine)
        self.table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.setStyleSheet(
            "QTableWidget {"
            "  background: #0D1117; border: none;"
            "  gridline-color: #21262D; outline: none;"
            "  color: #E6EDF3;"
            "  alternate-background-color: transparent; }"
            "QTableWidget::item { padding: 2px 6px;"
            "  border-right: 1px solid #21262D;"
            "  border-bottom: 1px solid #21262D; }"
            "QTableWidget::item:selected {"
            "  background-color: rgba(56,139,253,0.30); color: #E6EDF3; }"
            "QHeaderView { background: transparent; border: none; }"
            "QHeaderView::section {"
            "  background-color: #161B22; color: #8B949E;"
            "  padding: 10px 6px; border: none;"
            "  border-right: 1px solid #21262D;"
            "  border-bottom: 1px solid #21262D;"
            "  font-weight: 700; font-size: 10px;"
            "  letter-spacing: 0.5px; }"
        )

        # Stretch every column to fill the available width.
        header = self.table.horizontalHeader()
        for c in range(self.table.columnCount()):
            header.setSectionResizeMode(c, QHeaderView.ResizeMode.Stretch)
        header.setStretchLastSection(False)
        header.sectionResized.connect(
            lambda *_a: self._sync_day_header_widths()
        )

        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.table.setMinimumHeight(200)

        # Full-width container (no centering stretches) so the table
        # spans the same horizontal area as the filter card above.
        table_container = QHBoxLayout()
        table_container.addWidget(self.table)
        main_layout.addLayout(table_container, 1)

        # Pagination footer — Prev / numbered pages / Next + "Show N per page".
        pagination_layout = QHBoxLayout()
        pagination_layout.setSpacing(6)
        pagination_layout.setContentsMargins(0, 4, 0, 0)

        self._pager_btn_css = (
            "QPushButton { background: #161B22; border: 1px solid #30363D;"
            "  color: #C9D1D9; border-radius: 6px;"
            "  padding: 4px 10px; min-width: 28px; min-height: 26px;"
            "  font-size: 11px; font-weight: 700; }"
            "QPushButton:hover { background: rgba(255,255,255,0.05);"
            "  color: #E6EDF3; }"
            "QPushButton:disabled { color: #4d5560; border-color: #21262D; }"
        )
        self._pager_btn_active_css = (
            "QPushButton { background: transparent; border: 1px solid #388BFD;"
            "  color: #388BFD; border-radius: 6px;"
            "  padding: 4px 10px; min-width: 28px; min-height: 26px;"
            "  font-size: 11px; font-weight: 800; }"
        )

        self.btn_prev = QPushButton("‹ Prev")
        self.btn_prev.setCursor(Qt.PointingHandCursor)
        self.btn_prev.setStyleSheet(self._pager_btn_css)
        self.btn_prev.clicked.connect(self.prev_page)
        pagination_layout.addWidget(self.btn_prev)

        self._pager_buttons_row = QHBoxLayout()
        self._pager_buttons_row.setSpacing(4)
        pagination_layout.addLayout(self._pager_buttons_row)

        self.btn_next = QPushButton("Next ›")
        self.btn_next.setCursor(Qt.PointingHandCursor)
        self.btn_next.setStyleSheet(self._pager_btn_css)
        self.btn_next.clicked.connect(self.next_page)
        pagination_layout.addWidget(self.btn_next)

        pagination_layout.addStretch(1)

        ps_lbl = QLabel("Show")
        ps_lbl.setStyleSheet(
            "color: #8B949E; font-size: 11px; background: transparent;"
        )
        pagination_layout.addWidget(ps_lbl)
        self.page_size_combo = QComboBox()
        for n in (10, 25, 50, 100, 250):
            self.page_size_combo.addItem(str(n), n)
        idx = self.page_size_combo.findData(self.items_per_page)
        if idx >= 0:
            self.page_size_combo.setCurrentIndex(idx)
        self.page_size_combo.setFixedHeight(28)
        self.page_size_combo.setMinimumWidth(64)
        self.page_size_combo.setStyleSheet(
            "QComboBox { background: #161B22; border: 1px solid #30363D;"
            "  border-radius: 6px; padding: 2px 8px; color: #E6EDF3;"
            "  font-size: 11px; }"
        )
        self.page_size_combo.currentIndexChanged.connect(self._on_page_size_changed)
        pagination_layout.addWidget(self.page_size_combo)
        pagination_layout.addWidget(QLabel("per page"))

        # Legacy reference kept so existing setText callers don't crash.
        self.page_label = QLabel("Page 1 of 1"); self.page_label.hide()

        main_layout.addLayout(pagination_layout)

        self.setLayout(main_layout)

    def resizeEvent(self, event):
        """Hide Eff % column when width < 695px and move specific filter to new row at < 825px"""
        super().resizeEvent(event)
        width = event.size().width()
        try:
            self._sync_day_header_widths()
        except Exception:
            pass
        
        # Column widths: 90+100+80+70+90+50+50+60+70+20 = 680 (full)
        # Without Eff: 90+100+80+70+90+50+50+70+20 = 620
        
        # Hide Eff column on narrow viewports — the table itself stretches
        # to fill the available width (no setFixedWidth so it matches the
        # filter card's horizontal span).
        self.table.setColumnHidden(7, width < 695)
        
        # Legacy responsive logic (moving widgets between rows) removed —
        # the new two-row filter card already fits all widths. The
        # orphan QWidget containers used by the old layout (specific_row_widget,
        # filter_row3_widget) would otherwise pop up as top-level "python"
        # windows when shown.

    def load_regions_and_types(self):
        """Load unique regions and types for filters"""
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT DISTINCT region FROM cases ORDER BY region")
        regions = [row[0] for row in cursor.fetchall()]
        
        cursor.execute("SELECT DISTINCT tipo_caso FROM cases ORDER BY tipo_caso")
        types = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        
        self.filter_region.clear()
        self.filter_region.addItem("All")
        self.filter_region.addItems(regions)
        
        self.filter_type.clear()
        self.filter_type.addItem("All")
        self.filter_type.addItems(types)

    def load_all_cases(self):
        try:
            conn = get_connection()
            cursor = conn.cursor()
            # Load regular and OT cases together, marking source
            # Sort by date desc so OT and regular cases interleave by the
            # day they were registered — OT no longer drifts to the end
            # just because its ids live in a separate sequence. Use
            # hora_inicio + case_id as tie-breakers within a day.
            cursor.execute("""
                SELECT * FROM (
                    SELECT id, case_id, doctor, region, tipo_caso,
                           fecha, tiempo_real, std_time, efficiency, estado, case_value,
                           COALESCE(count_production, 1) as count_production,
                           'reg' as source, comments,
                           COALESCE(hora_inicio, '00:00') AS _ord_hh
                    FROM cases
                    UNION ALL
                    SELECT id, case_id, doctor, region, tipo_caso,
                           fecha, tiempo_real, std_time, efficiency, estado, case_value,
                           COALESCE(count_production, 1) as count_production,
                           'ot' as source, comments,
                           COALESCE(hora_inicio, '00:00') AS _ord_hh
                    FROM ot_cases
                )
                ORDER BY fecha DESC, _ord_hh DESC, case_id DESC
            """)
            self.all_cases = cursor.fetchall()
            conn.close()
        except Exception as exc:
            # Keep prior snapshot if DB momentarily locked
            print(f"[HistoryTab] load_all_cases failed: {exc}")
            return
        self.load_regions_and_types()
        self.filter_cases()

    def _open_filter_popup(self):
        """Popup with Region / Type / Count combos anchored under the
        Filters button. Apply writes into the hidden combos which the
        existing filter pipeline reads from."""
        from PySide6.QtWidgets import QFrame as _QFp, QComboBox as _QCp
        try:
            from .widgets import _icon_url as _icu_p
            _chev_p = _icu_p("tabler_chevron_down.svg")
        except Exception:
            _chev_p = ""

        popup = _QFp(self, Qt.Popup)
        popup.setObjectName("filterPopup")
        popup.setStyleSheet(
            "#filterPopup { background-color: #161B22; border: 1px solid #30363D;"
            "  border-radius: 10px; }"
            "QLabel { color: #8B949E; font-size: 10px; font-weight: 700;"
            "  background: transparent; padding: 0; }"
            "QComboBox { background: #0D1117; border: 1px solid #30363D;"
            "  border-radius: 6px; padding: 4px 22px 4px 8px; color: #E6EDF3;"
            "  font-size: 11px; min-height: 26px; }"
            "QComboBox::drop-down { subcontrol-origin: padding;"
            "  subcontrol-position: right center; width: 22px; border: none; }"
            f"QComboBox::down-arrow {{ image: url({_chev_p});"
            "  width: 12px; height: 12px; }"
            "QPushButton { border-radius: 6px; padding: 6px 12px;"
            "  font-size: 11px; font-weight: 700; }"
            "QPushButton#apply { background: #1757D4; border: 1px solid #1757D4;"
            "  color: white; }"
            "QPushButton#apply:hover { background: #1F6FEB; }"
            "QPushButton#clear { background: transparent; border: 1px solid #30363D;"
            "  color: #E6EDF3; }"
            "QPushButton#clear:hover { background: rgba(255,255,255,0.05); }"
        )
        lay = QVBoxLayout(popup)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(6)

        # Build Region/Type lists constrained by the OTHER active
        # filters (date + status + search + doctor) so the popup only
        # shows values that exist in the currently displayed rows.
        where = []
        params: list = []
        if self.specific_date_check.isChecked():
            where.append("fecha = ?")
            params.append(self.specific_date.date().toString("yyyy-MM-dd"))
        st_sel = self.status_filter.currentText() if hasattr(self, "status_filter") else "All"
        if st_sel and st_sel != "All":
            where.append("estado = ?"); params.append(st_sel)
        s_txt = (self.search_input.text() or "").strip()
        if s_txt:
            where.append("(case_id LIKE ? OR doctor LIKE ?)")
            params += [f"%{s_txt}%", f"%{s_txt}%"]
        d_txt = (self.filter_doctor.text() or "").strip()
        if d_txt:
            where.append("doctor LIKE ?"); params.append(f"%{d_txt}%")
        clause = (" WHERE " + " AND ".join(where)) if where else ""

        try:
            conn = get_connection()
            cur = conn.cursor()
            cur.execute(
                f"SELECT DISTINCT region FROM cases{clause} ORDER BY region",
                params,
            )
            regions = [r[0] for r in cur.fetchall() if r[0]]
            cur.execute(
                f"SELECT DISTINCT tipo_caso FROM cases{clause}"
                " ORDER BY tipo_caso", params,
            )
            types = [r[0] for r in cur.fetchall() if r[0]]
            conn.close()
        except Exception:
            regions, types = [], []

        def _combo(values, current, all_label="All"):
            c = _QCp()
            c.addItem(all_label)
            for v in values:
                c.addItem(str(v))
            if current and current in (all_label, *values):
                c.setCurrentText(current)
            return c

        lay.addWidget(QLabel("REGION"))
        region_c = _combo(regions, self.filter_region.currentText())
        lay.addWidget(region_c)

        lay.addSpacing(2)
        lay.addWidget(QLabel("TYPE"))
        type_c = _combo(types, self.filter_type.currentText())
        lay.addWidget(type_c)


        lay.addSpacing(8)
        btn_row = QHBoxLayout(); btn_row.setSpacing(8)
        clear_btn = QPushButton("Clear"); clear_btn.setObjectName("clear")
        apply_btn = QPushButton("Apply"); apply_btn.setObjectName("apply")
        btn_row.addWidget(clear_btn); btn_row.addStretch(); btn_row.addWidget(apply_btn)
        lay.addLayout(btn_row)

        def _apply():
            self.filter_region.blockSignals(True)
            self.filter_region.setCurrentText(region_c.currentText())
            self.filter_region.blockSignals(False)
            self.filter_type.blockSignals(True)
            self.filter_type.setCurrentText(type_c.currentText())
            self.filter_type.blockSignals(False)
            active = (
                (region_c.currentText() not in ("", "All"))
                or (type_c.currentText() not in ("", "All"))
            )
            self._filters_btn.setText("  Filters •" if active else "  Filters")
            self.on_filter_changed()
            popup.close()

        def _clear():
            for c in (self.filter_region, self.filter_type):
                c.blockSignals(True)
                c.setCurrentIndex(0)
                c.blockSignals(False)
            self._filters_btn.setText("  Filters")
            self.on_filter_changed()
            popup.close()

        apply_btn.clicked.connect(_apply)
        clear_btn.clicked.connect(_clear)

        btn = self._filters_btn
        pos = btn.mapToGlobal(btn.rect().bottomRight())
        popup.adjustSize()
        popup.move(pos.x() - popup.width(), pos.y() + 6)
        popup.show()

    def on_filter_changed(self):
        """Reset to page 1 when filter changes"""
        self.current_page = 1
        self.filter_cases()
    
    def on_specific_date_toggled(self, state):
        """The date picker stays editable at all times. The checkbox only
        toggles whether the chosen date filters the table — re-render
        the list once the state flips."""
        self.on_filter_changed()

    def prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self.filter_cases()

    def next_page(self):
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.filter_cases()

    def filter_cases(self):
        search_text = self.search_input.text().lower()
        status_filter = self.status_filter.currentText()
        
        # Check if specific date is enabled
        use_specific_date = self.specific_date_check.isChecked()
        if use_specific_date:
            specific = self.specific_date.date().toString("yyyy-MM-dd")
            date_from = specific
            date_to = specific
        else:
            date_from = self.date_from.date().toString("yyyy-MM-dd")
            date_to = self.date_to.date().toString("yyyy-MM-dd")
        
        region_filter = self.filter_region.currentText()
        type_filter = self.filter_type.currentText()
        count_filter = self.count_filter.currentText()
        doctor_filter = self.filter_doctor.text().lower()

        filtered = []
        for case in self.all_cases:
            # case: id, case_id, doctor, region, tipo_caso, fecha, tiempo_real, std_time, efficiency, estado, case_value
            if search_text:
                cid = str(case[1] or "").lower()
                doc = str(case[2] or "").lower()
                if search_text not in cid and search_text not in doc:
                    continue
            if status_filter != "All" and status_filter != case[9]:
                continue
            if case[5] < date_from or case[5] > date_to:
                continue
            if region_filter != "All" and case[3] != region_filter:
                continue
            if type_filter != "All" and case[4] != type_filter:
                continue
            counts_for_production = case[11] if (len(case) > 11 and case[11] is not None) else 1
            if count_filter == "Counted" and counts_for_production == 0:
                continue
            if count_filter == "NC" and counts_for_production != 0:
                continue
            # Comments toggle: "with" → only commented, "without" → only blank.
            cmts_state = getattr(self, "_filter_comments", "") or ""
            if cmts_state:
                # comments at index 13 (... source, comments)
                row_cmts = (case[13] if len(case) > 13 else "") or ""
                has_cmts = bool(str(row_cmts).strip())
                if cmts_state == "with" and not has_cmts:
                    continue
                if cmts_state == "without" and has_cmts:
                    continue
            # Counted toggle: 'counted' / 'nc' / '' (uses same int flag).
            cnt_state = getattr(self, "_filter_counted", "") or ""
            if cnt_state == "counted" and counts_for_production == 0:
                continue
            if cnt_state == "nc" and counts_for_production != 0:
                continue
            if doctor_filter and doctor_filter not in (case[2] or "").lower():
                continue
            filtered.append(case)

        self.filtered_cases = filtered
        
        # Calculate pagination
        total_items = len(filtered)
        self.total_pages = max(1, (total_items + self.items_per_page - 1) // self.items_per_page)
        
        if self.current_page > self.total_pages:
            self.current_page = self.total_pages
        if self.current_page < 1:
            self.current_page = 1
        
        self.page_label.setText(f"Page {self.current_page} of {self.total_pages}")
        self._rebuild_pager()
        self.btn_prev.setEnabled(self.current_page > 1)
        self.btn_next.setEnabled(self.current_page < self.total_pages)
        
        # Get items for current page
        start_idx = (self.current_page - 1) * self.items_per_page
        end_idx = start_idx + self.items_per_page
        page_items = filtered[start_idx:end_idx]

        # Flat row layout — no day-group headers in History.
        self.table.setRowCount(len(page_items))
        self.table.clearSpans()
        self._day_hdr_widgets = []

        # Update summary stats for ALL filtered items (not only current
        # page). Non-counting cases (count_production == 0) are excluded
        # from every KPI so totals only reflect production-counting work.
        counted = [
            c for c in filtered
            if (c[11] if len(c) > 11 and c[11] is not None else 1) != 0
        ]
        total_cases = len(counted)
        total_time = sum((c[6] or 0) for c in counted)
        total_value = sum((c[10] or 0) for c in counted)
        total_ue = sum(
            calculate_equivalent_units(
                self.units_eq,
                c[3],
                c[4],
                (c[10] or 0),
                count=1,
            )
            for c in counted
        )

        # Calculate downtime credit for the filtered date range
        use_specific = self.specific_date_check.isChecked()
        if use_specific:
            dt_from = self.specific_date.date().toString("yyyy-MM-dd")
            dt_to = dt_from
        else:
            dt_from = self.date_from.date().toString("yyyy-MM-dd")
            dt_to = self.date_to.date().toString("yyyy-MM-dd")

        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT SUM(duracion) FROM downtimes "
            "WHERE fecha >= ? AND fecha <= ? AND (status = 'approved' OR status IS NULL)",
            (dt_from, dt_to),
        )
        dt_row = cur.fetchone()
        conn.close()
        total_downtime_mins = dt_row[0] if dt_row and dt_row[0] else 0.0
        downtime_value = (total_downtime_mins / DAILY_BASE_MINUTES) * 100 if total_downtime_mins > 0 else 0

        combined_value = total_value + downtime_value

        if total_downtime_mins > 0:
            self.stats_label.setText(
                f"Total: {total_cases} | Time: {total_time:.0f}m | "
                f"Production: {combined_value:.2f}% (Cases: {total_value:.2f}% + Downtime: {downtime_value:.2f}%)"
            )
            self.stats_label2.setText(f"Equivalent Units: {total_ue:.2f}")
        else:
            self.stats_label.setText(
                f"Total: {total_cases} | Time: {total_time:.0f}m | Value: {total_value:.2f}% | UE: {total_ue:.2f}"
            )

        # New KPI card values.
        if hasattr(self, "_kpi_cases"):
            self._kpi_cases.setText(f"{total_cases:,}")
            self._kpi_time.setText(f"{total_time:,.0f}m")
            self._kpi_value.setText(f"{(combined_value if total_downtime_mins > 0 else total_value):,.2f}%")
            self._kpi_ue.setText(f"{total_ue:,.2f}")

        is_light = False
        try:
            from qfluentwidgets.common.style_sheet import isDarkTheme
            is_light = not isDarkTheme()
        except Exception:
            try:
                is_light = "#F6F8FA" in (QApplication.instance().styleSheet() or "")
            except Exception:
                is_light = False
        light_colors = get_light_theme_colors()

        for idx, case in enumerate(page_items):
            row_idx = idx
            # Match Production tab row height so the table reads the same.
            self.table.setRowHeight(row_idx, 32)
            bg_color = light_row_bg(idx, light_colors) if is_light else (QColor(43, 43, 43) if (idx % 2 == 0) else QColor(55, 55, 55))
            # If this is an OT case (source at index 12), use a distinct blue tint
            try:
                source = case[12]
            except Exception:
                source = 'reg'
            if source == 'ot':
                if is_light:
                    bg_color = QColor(mix_hex(light_colors.get("selection_bg", "#DDF4FF"), "#FFFFFF", 0.18 if (idx % 2 == 0) else 0.30))
                else:
                    # Stronger blue tint so OT rows are clearly visible in dark mode.
                    bg_color = QColor(34, 62, 122) if (idx % 2 == 0) else QColor(44, 72, 138)
            # Non-counting cases — soft orange tint (matches Production tab).
            try:
                _cp = case[11] if len(case) > 11 and case[11] is not None else 1
            except Exception:
                _cp = 1
            if _cp == 0:
                bg_color = QColor("#FFE9CC") if is_light else QColor("#3A2E1F")
            bg_brush = QBrush(bg_color)
            
            # Case ID — bold, with comment chat icon (clickable) when
            # there's a comment attached. The QTableWidgetItem stays as
            # the raw case_id so selection/sort logic still works; the
            # visible cell is overlaid via setCellWidget.
            counts_for_production = case[11] if (len(case) > 11 and case[11] is not None) else 1
            comment = (case[13] if len(case) > 13 else "") or ""

            suffix_text = ""
            if source == 'ot':
                suffix_text += " (OT)"
            if counts_for_production == 0:
                suffix_text += " (NC)"

            # Underlying item — empty display text so it doesn't bleed
            # through the cell widget; the raw case_id is stashed in
            # UserRole for backend readers.
            case_id_item = QTableWidgetItem("")
            case_id_item.setData(Qt.ItemDataRole.UserRole, str(case[1]))
            case_id_item.setBackground(bg_brush)
            case_id_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            bold_font = QFont(); bold_font.setBold(True)
            if counts_for_production == 0:
                bold_font.setItalic(True)
            case_id_item.setFont(bold_font)
            if counts_for_production == 0:
                fg = QColor("#A15C00") if is_light else QColor("#F0883E")
            elif source == 'ot':
                fg = QColor(mix_hex(light_colors.get("accent", "#0969DA"), "#000000", 0.15)) if is_light else QColor("#7DB3FF")
            else:
                fg = CLR_FG_DARK if is_light else CLR_FG_LIGHT
            case_id_item.setForeground(QBrush(fg))
            if comment.strip():
                case_id_item.setToolTip(comment.strip())
            self.table.setItem(row_idx, 0, case_id_item)

            # Overlay a cell widget: case id text + small message-circle
            # icon that opens a popup with the comment when clicked.
            self.table.setCellWidget(
                row_idx, 0,
                self._build_case_id_cell(
                    case_id=str(case[1]),
                    suffix=suffix_text,
                    comment=comment.strip(),
                    text_color=fg.name(),
                    bold=True,
                    italic=(counts_for_production == 0),
                ),
            )
            
            # Doctor - Bold
            doctor_item = QTableWidgetItem(str(case[2] or "-"))
            doctor_item.setBackground(bg_brush)
            doctor_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            doctor_item.setFont(bold_font)
            doctor_item.setForeground(QBrush(CLR_FG_DARK if is_light else CLR_FG_LIGHT))
            self.table.setItem(row_idx, 1, doctor_item)
            
            # Region
            region_item = QTableWidgetItem(str(case[3]))
            region_item.setBackground(bg_brush)
            region_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            region_item.setForeground(QBrush(CLR_FG_DARK if is_light else CLR_FG_LIGHT))
            self.table.setItem(row_idx, 2, region_item)
            
            # Type
            type_item = QTableWidgetItem(str(case[4]))
            type_item.setBackground(bg_brush)
            type_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            type_item.setForeground(QBrush(CLR_FG_DARK if is_light else CLR_FG_LIGHT))
            self.table.setItem(row_idx, 3, type_item)
            
            # Date
            date_item = QTableWidgetItem(str(case[5]))
            date_item.setBackground(bg_brush)
            date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            date_item.setForeground(QBrush(CLR_FG_DARK if is_light else CLR_FG_LIGHT))
            self.table.setItem(row_idx, 4, date_item)
            
            # Time — colour the TEXT only (cell bg neutral) to match the
            # Production tab style.
            tiempo_real = case[6] or 0
            estado = case[9]
            time_item = QTableWidgetItem(f"{tiempo_real:.0f}")
            time_item.setBackground(bg_brush)
            if estado == "OK":
                time_item.setForeground(QBrush(QColor("#3FB950")))
            else:
                time_item.setForeground(QBrush(QColor("#F85149")))
            time_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 5, time_item)
            
            # Std
            std_item = QTableWidgetItem(f"{case[7]:.1f}")
            std_item.setBackground(bg_brush)
            std_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            std_item.setForeground(QBrush(CLR_FG_DARK if is_light else CLR_FG_LIGHT))
            self.table.setItem(row_idx, 6, std_item)
            
            # Efficiency — text-coloured (green/amber/red) on neutral bg.
            try:
                eff_val = float(case[8])
            except Exception:
                eff_val = None
            efficiency_item = QTableWidgetItem(f"{case[8]:.0f}%")
            efficiency_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            efficiency_item.setBackground(bg_brush)
            if eff_val is not None:
                if eff_val >= 100:
                    efficiency_item.setForeground(QBrush(QColor("#3FB950")))
                elif eff_val >= 95:
                    efficiency_item.setForeground(QBrush(QColor("#D29922")))
                else:
                    efficiency_item.setForeground(QBrush(QColor("#F85149")))
            else:
                col = "#3FB950" if estado == "OK" else "#F85149"
                efficiency_item.setForeground(QBrush(QColor(col)))
            self.table.setItem(row_idx, 7, efficiency_item)
            
            # Case Value
            value_item = QTableWidgetItem(f"{case[10]:.2f}%")
            value_item.setBackground(bg_brush)
            value_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            value_item.setForeground(QBrush(CLR_FG_DARK if is_light else CLR_FG_LIGHT))
            self.table.setItem(row_idx, 8, value_item)

    def _on_page_size_changed(self, _idx: int):
        try:
            self.items_per_page = int(self.page_size_combo.currentData())
        except Exception:
            return
        self.current_page = 1
        self.filter_cases()

    def _rebuild_pager(self):
        row = self._pager_buttons_row
        while row.count():
            item = row.takeAt(0)
            w = item.widget()
            if w is not None:
                w.setParent(None)
        total = max(1, int(getattr(self, "total_pages", 1) or 1))
        cur = max(1, int(getattr(self, "current_page", 1) or 1))

        def _btn(label, page=None, active=False):
            b = QPushButton(str(label))
            b.setCursor(Qt.PointingHandCursor)
            b.setStyleSheet(
                self._pager_btn_active_css if active else self._pager_btn_css
            )
            if page is not None:
                b.clicked.connect(lambda _=False, p=page: self._goto_page(p))
            return b

        pages: list = []
        if total <= 7:
            pages = list(range(1, total + 1))
        else:
            pages.append(1)
            if cur > 3:
                pages.append("…")
            for p in range(max(2, cur - 1), min(total - 1, cur + 1) + 1):
                pages.append(p)
            if cur < total - 2:
                pages.append("…")
            pages.append(total)
        for p in pages:
            if p == "…":
                lbl = QLabel("…")
                lbl.setStyleSheet(
                    "color: #8B949E; font-size: 12px; padding: 0 6px;"
                    " background: transparent;"
                )
                row.addWidget(lbl)
            else:
                row.addWidget(_btn(p, page=p, active=(p == cur)))

    def _goto_page(self, page: int):
        page = max(1, min(int(page), int(getattr(self, "total_pages", 1) or 1)))
        if page == self.current_page:
            return
        self.current_page = page
        self.filter_cases()

    def _build_case_id_cell(self, *, case_id, suffix, comment,
                             text_color, bold, italic):
        """Cell widget for column 0 — both children vertically centered;
        no slot reserved when there's no comment so the bare case_id is
        centred on the cell."""
        from PySide6.QtWidgets import (
            QWidget as _QW, QHBoxLayout as _QH, QLabel as _QL,
            QToolButton as _QTB, QSizePolicy as _SP,
        )
        wrap = _QW()
        wrap.setStyleSheet("background: transparent; border: none;")
        wrap.setSizePolicy(_SP.Policy.Expanding, _SP.Policy.Expanding)
        h = _QH(wrap)
        h.setContentsMargins(0, 0, 0, 0)
        h.setSpacing(4)
        h.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        h.addStretch(1)

        lbl = _QL(f"{case_id}{suffix}")
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        f = QFont(); f.setBold(bold); f.setItalic(italic)
        lbl.setFont(f)
        lbl.setStyleSheet(
            f"color: {text_color}; background: transparent;"
        )

        if comment:
            try:
                from .tabler_icons import TablerIcon as _TI_cm
                from PySide6.QtCore import QSize as _QS_cm
                btn = _QTB()
                btn.setCursor(Qt.PointingHandCursor)
                btn.setFixedSize(20, 20)
                btn.setIcon(
                    _TI_cm("tabler_message_circle.svg").icon(color=QColor("#58A6FF"))
                )
                btn.setIconSize(_QS_cm(14, 14))
                btn.setToolTip(comment)
                btn.setStyleSheet(
                    "QToolButton { background: transparent; border: none; }"
                    "QToolButton:hover { background: rgba(56,139,253,0.12);"
                    "  border-radius: 4px; }"
                )
                btn.clicked.connect(
                    lambda _=False, cid=case_id, c=comment:
                        self._show_comment_popup(cid, c)
                )
                h.addWidget(btn, 0, Qt.AlignmentFlag.AlignVCenter)
            except Exception:
                pass

        h.addWidget(lbl, 0, Qt.AlignmentFlag.AlignVCenter)
        h.addStretch(1)
        return wrap

    def _show_comment_popup(self, case_id: str, comment: str):
        """Fluent modal showing the full comment text for a case."""
        try:
            from qfluentwidgets import MessageBoxBase
            from .tabler_icons import TablerIcon
            from PySide6.QtWidgets import (
                QFrame as _QF, QTextEdit as _QTE, QToolButton as _QTB,
            )
            from PySide6.QtCore import QSize as _QSz
        except Exception:
            QMessageBox.information(self, f"Comment · {case_id}", comment)
            return

        host = self

        class _Sheet(MessageBoxBase):
            def __init__(_s, h):
                super().__init__(h.window() if h is not None else None)
                try:
                    _s.setMaskColor(QColor(0, 0, 0, 170))
                except Exception:
                    pass
                _s.widget.setObjectName("cmtCard")
                apply_fluent_modal_palette(_s, "cmtCard")
                _s.viewLayout.setContentsMargins(22, 18, 22, 12)
                _s.viewLayout.setSpacing(12)

                hdr = QHBoxLayout(); hdr.setSpacing(12)
                ic = _QTB(); ic.setEnabled(False)
                ic.setIcon(TablerIcon("tabler_message_circle.svg").icon(color=QColor("#58A6FF")))
                ic.setIconSize(_QSz(22, 22))
                ic.setStyleSheet(
                    "background: rgba(56,139,253,0.14); border: none;"
                    " border-radius: 10px; padding: 6px;"
                )
                tc = QVBoxLayout(); tc.setSpacing(2)
                t = QLabel("Comment")
                t.setStyleSheet(
                    "color: #E6EDF3; font-size: 15px; font-weight: 700;"
                    " background: transparent;"
                )
                s = QLabel(f"Case {case_id}")
                s.setStyleSheet(
                    "color: #8B949E; font-size: 11px; background: transparent;"
                )
                tc.addWidget(t); tc.addWidget(s)
                hdr.addWidget(ic, 0, Qt.AlignmentFlag.AlignTop)
                hdr.addLayout(tc, 1)
                _s.viewLayout.addLayout(hdr)

                box = _QTE()
                box.setReadOnly(True)
                box.setPlainText(comment)
                box.setStyleSheet(
                    "QTextEdit { background: #161B22; border: 1px solid #30363D;"
                    "  border-radius: 6px; padding: 8px 10px; color: #E6EDF3;"
                    "  font-size: 12px; }"
                )
                box.setMinimumHeight(120)
                _s.viewLayout.addWidget(box)
                _s.widget.setMinimumWidth(460)

                _s.cancelButton.hide()
                _s.yesButton.setText("Close")
                _s.yesButton.setFixedWidth(120)
                _s.yesButton.setStyleSheet(
                    "QPushButton { background: #1e63e4; border: 1px solid #1e63e4;"
                    "  color: white; border-radius: 6px; padding: 8px 22px;"
                    "  font-weight: 700; font-size: 12px; }"
                    "QPushButton:hover { background: #2a73f3; }"
                )

        _Sheet(host).exec()

    def _build_day_header_widget(self, *, fecha, cases_count, value,
                                  units, time_min, breakdown_text):
        from PySide6.QtWidgets import QFrame as _QF_dh, QSizePolicy as _SP_dh
        wrap = _QF_dh()
        wrap.setObjectName("dayBand")
        wrap.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        wrap.setAutoFillBackground(True)
        wrap.setSizePolicy(_SP_dh.Policy.Expanding, _SP_dh.Policy.Expanding)
        wrap.setStyleSheet(
            "#dayBand { background-color: #15233D; border: none;"
            "  border-bottom: 1px solid #0D1117; }"
            "QLabel { background: transparent; color: #E6EDF3;"
            "  font-size: 11px; }"
        )
        v = QVBoxLayout(wrap)
        v.setContentsMargins(14, 4, 14, 4)
        v.setSpacing(2)
        line1 = (
            f"   📅  {fecha}     {cases_count} cases     "
            f"Value: {value:.2f}%     "
            f"Units: {units:.2f}     Time: {time_min:.0f}m"
        )
        top_lbl = QLabel(line1)
        tf = QFont(); tf.setBold(True); tf.setPointSize(10)
        top_lbl.setFont(tf)
        top_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        top_lbl.setStyleSheet("color: #E6EDF3; background: transparent;")
        v.addWidget(top_lbl)
        if breakdown_text:
            sub = QLabel(breakdown_text)
            sf = QFont(); sf.setPointSize(8)
            sub.setFont(sf)
            sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
            sub.setStyleSheet("color: #B0B8C0; background: transparent;")
            v.addWidget(sub)
        return wrap

    def _sync_day_header_widths(self):
        widgets = getattr(self, "_day_hdr_widgets", None)
        if not widgets:
            return
        total = self.table.viewport().width()
        if total <= 0:
            for c in range(self.table.columnCount()):
                if not self.table.isColumnHidden(c):
                    total += self.table.columnWidth(c)
        if total <= 0:
            return
        for row_idx, w in widgets:
            try:
                w.setMinimumWidth(total)
                w.setMaximumWidth(total)
                y = self.table.rowViewportPosition(row_idx)
                h = self.table.rowHeight(row_idx)
                w.setGeometry(0, y, total, h)
            except Exception:
                pass

    def export_csv(self):
        """Export to Excel with native table filters, or CSV as fallback"""
        if OPENPYXL_AVAILABLE:
            self.export_excel()
        else:
            self.export_csv_simple()
    
    def export_excel(self):
        """Export to Excel with native table filters and Summary sheet."""
        _ueq = self.units_eq
        COLUMNS = [
            ("Case ID",      lambda c: c[1]),
            ("Doctor",       lambda c: c[2] or "-"),
            ("Region",       lambda c: c[3]),
            ("Type",         lambda c: c[4]),
            ("Date",         lambda c: c[5]),
            ("Time (min)",   lambda c: round(c[6], 1)),
            ("Std (min)",    lambda c: round(c[7], 1)),
            ("Efficiency %", lambda c: round(c[8], 1)),
            ("Status",       lambda c: c[9]),
            ("Case Value %", lambda c: round(c[10], 2)),
            ("UE",           lambda c, _u=_ueq: round(calculate_equivalent_units(_u, c[3], c[4], (c[10] or 0), count=1), 2)),
        ]
        DATA_START = 1  # table header row (row 1)

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Report", "", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return
        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'

        export_cases = self.filtered_cases

        try:
            wb = Workbook()

            # ── Sheet 1: Cases Data ─────────────────────────────────────────────
            ws_data = wb.active
            ws_data.title = "Cases Data"

            # Styles
            ok_fill     = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            low_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            hdr_fill    = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'),  bottom=Side(style='thin'),
            )
            center  = Alignment(horizontal="center", vertical="center")
            ncols   = len(COLUMNS)
            last_col = get_column_letter(ncols)

            # ── Row 1: Table headers ──────────────────────────────────────────
            ws_data.row_dimensions[DATA_START].height = 18
            for col_pos, (label, _) in enumerate(COLUMNS, 1):
                cell = ws_data.cell(DATA_START, col_pos, label)
                cell.font      = header_font
                cell.fill      = hdr_fill
                cell.alignment = center
                cell.border    = thin_border

            # ── Rows 2+: Data ─────────────────────────────────────────────────
            for row_idx, case in enumerate(export_cases, DATA_START + 1):
                row_fill = ok_fill if case[9] == "OK" else low_fill
                for col_pos, (_, extractor) in enumerate(COLUMNS, 1):
                    cell = ws_data.cell(row_idx, col_pos, extractor(case))
                    cell.alignment = center
                    cell.fill      = row_fill
                    cell.border    = thin_border

            # ── Excel Table (with native filter arrows) ────────────────────────
            last_data_row = max(DATA_START + len(export_cases), DATA_START + 1)
            tbl_ref = f"A{DATA_START}:{last_col}{last_data_row}"
            tbl = Table(displayName="CasesData", ref=tbl_ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True,
                showFirstColumn=False,
                showLastColumn=False,
                showColumnStripes=False,
            )
            ws_data.add_table(tbl)

            # Freeze header row
            ws_data.freeze_panes = "A2"

            # ── Auto-fit column widths ────────────────────────────────────────
            for ci in range(1, ncols + 1):
                col_letter = get_column_letter(ci)
                max_len = 10
                for ri in range(DATA_START, last_data_row + 1):
                    try:
                        val = ws_data.cell(ri, ci).value
                        cl = len(str(val)) if val is not None else 0
                        if cl > max_len:
                            max_len = cl
                    except Exception:
                        pass
                ws_data.column_dimensions[col_letter].width = max_len + 3
            
            # ===== SHEET 2: SUMMARY & CHARTS =====
            ws_charts = wb.create_sheet("Summary & Charts")
            
            # Calculate summary statistics
            total_cases = len(export_cases)
            ok_cases = sum(1 for c in export_cases if c[9] == "OK")
            low_cases = total_cases - ok_cases
            avg_efficiency = sum(c[8] for c in export_cases) / total_cases if total_cases > 0 else 0
            total_value = sum(c[10] for c in export_cases)
            
            # Summary section
            ws_charts['A1'] = "PRODUCTION SUMMARY REPORT"
            ws_charts['A1'].font = Font(bold=True, size=16, color="4A90D9")
            ws_charts.merge_cells('A1:D1')
            
            summary_data = [
                ("Total Cases:", total_cases),
                ("OK Cases:", ok_cases),
                ("LOW Cases:", low_cases),
                ("OK Rate:", f"{(ok_cases/total_cases*100):.1f}%" if total_cases > 0 else "0%"),
                ("Average Efficiency:", f"{avg_efficiency:.1f}%"),
                ("Total Value:", f"{total_value:.2f}%"),
            ]
            
            for row, (label, value) in enumerate(summary_data, 3):
                ws_charts.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws_charts.cell(row=row, column=2, value=value)
            
            # Production by Region
            region_stats = {}
            for case in export_cases:
                region = case[3]
                if region not in region_stats:
                    region_stats[region] = {"count": 0, "value": 0, "efficiency": []}
                region_stats[region]["count"] += 1
                region_stats[region]["value"] += case[10]
                region_stats[region]["efficiency"].append(case[8])
            
            # Region table for chart
            ws_charts['A12'] = "Production by Region"
            ws_charts['A12'].font = Font(bold=True, size=12)
            ws_charts['A13'] = "Region"
            ws_charts['B13'] = "Cases"
            ws_charts['C13'] = "Total Value %"
            ws_charts['A13'].font = Font(bold=True)
            ws_charts['B13'].font = Font(bold=True)
            ws_charts['C13'].font = Font(bold=True)
            
            row_num = 14
            for region, stats in sorted(region_stats.items()):
                ws_charts.cell(row=row_num, column=1, value=region)
                ws_charts.cell(row=row_num, column=2, value=stats["count"])
                ws_charts.cell(row=row_num, column=3, value=round(stats["value"], 2))
                row_num += 1
            
            # Bar Chart - Cases by Region
            if len(region_stats) > 0:
                chart1 = BarChart()
                chart1.type = "col"
                chart1.style = 10
                chart1.title = "Cases by Region"
                chart1.y_axis.title = "Cases"
                chart1.x_axis.title = "Region"
                
                data = Reference(ws_charts, min_col=2, min_row=13, max_row=13 + len(region_stats), max_col=2)
                cats = Reference(ws_charts, min_col=1, min_row=14, max_row=13 + len(region_stats))
                chart1.add_data(data, titles_from_data=True)
                chart1.set_categories(cats)
                chart1.shape = 4
                chart1.width = 12
                chart1.height = 8
                ws_charts.add_chart(chart1, "E12")
            
            # Status Pie Chart data
            ws_charts['A' + str(row_num + 2)] = "Status Distribution"
            ws_charts['A' + str(row_num + 2)].font = Font(bold=True, size=12)
            ws_charts['A' + str(row_num + 3)] = "Status"
            ws_charts['B' + str(row_num + 3)] = "Count"
            ws_charts['A' + str(row_num + 3)].font = Font(bold=True)
            ws_charts['B' + str(row_num + 3)].font = Font(bold=True)
            ws_charts['A' + str(row_num + 4)] = "OK"
            ws_charts['B' + str(row_num + 4)] = ok_cases
            ws_charts['A' + str(row_num + 5)] = "LOW"
            ws_charts['B' + str(row_num + 5)] = low_cases
            
            # Pie Chart - Status Distribution
            if total_cases > 0:
                pie = PieChart()
                pie.title = "Status Distribution"
                labels = Reference(ws_charts, min_col=1, min_row=row_num + 4, max_row=row_num + 5)
                data = Reference(ws_charts, min_col=2, min_row=row_num + 3, max_row=row_num + 5)
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                pie.width = 10
                pie.height = 8
                ws_charts.add_chart(pie, "E" + str(row_num + 2))
            
            # Save file
            wb.save(file_path)
            QMessageBox.information(self, "Export Successful", 
                f"Report exported to:\n{file_path}\n\nIncludes:\n• Formatted data table\n• Summary statistics\n• Charts by region and status")
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Error exporting file:\n{str(e)}")
    
    def export_csv_simple(self):
        """Simple CSV export as fallback"""
        file_path, _ = QFileDialog.getSaveFileName(self, "Export History", "", "CSV Files (*.csv)")
        if file_path:
            try:
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow([
                        "Case ID", "Doctor", "Region", "Type",
                        "Date", "Time (min)", "Std (min)", "Efficiency (%)", "Status", "Case Value (%)"
                    ])
                    for case in self.filtered_cases:
                        writer.writerow([
                            case[1], case[2], case[3], case[4],
                            case[5], f"{case[6]:.1f}", f"{case[7]:.1f}", 
                            f"{case[8]:.1f}", case[9], f"{case[10]:.3f}"
                        ])
                QMessageBox.information(self, "Export Successful", f"CSV exported to:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Export Error", f"Error: {str(e)}")

    def update_font_sizes(self, _new_size: int = 0):
        """Re-render the table so QFont calls inside load_all_cases pick up
        the new global scale."""
        try:
            self.load_all_cases()
        except Exception:
            pass

    def update_theme_labels(self, is_light: bool):
        """Apply theme colors; in light mode use user-selected palette."""
        colors = get_light_theme_colors()
        light_table_css = (
            f' QTableWidget {{ background-color: {colors["surface_bg"]}; gridline-color: {colors["border"]}; border: 1px solid {colors["border"]}; }} '
            f' QHeaderView::section {{ background-color: {light_header_bg(colors)}; color: {light_header_fg(colors)}; border: 1px solid {colors["border"]}; padding: 4px; }} '
        )

        apply_table_theme(
            self,
            is_light,
            light_append_css=light_table_css,
            adaptive_fg_by_bg=True,
            adaptive_default_fg=CLR_FG_DARK if is_light else CLR_FG_LIGHT,
        )

        # Title color: make History title dark in light mode, keep blue in dark
        try:
            for lbl in self.findChildren(QLabel):
                if lbl.text().strip() == "Case History":
                    if is_light:
                        lbl.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {colors['text_primary']};")
                    else:
                        lbl.setStyleSheet("font-size: 16px; font-weight: bold; color: #4aa3ff;")
                    break
        except Exception:
            pass

        # Stats label (Total | Time | Value) — use dark text in light mode
        try:
            if hasattr(self, 'stats_label') and self.stats_label:
                if is_light:
                    self.stats_label.setStyleSheet(f"font-size: 12px; color: {colors['text_primary']}; font-weight: bold;")
                else:
                    self.stats_label.setStyleSheet("font-size: 12px; color: #9CC3FF; font-weight: bold;")
            if hasattr(self, 'stats_label2') and self.stats_label2:
                if is_light:
                    self.stats_label2.setStyleSheet(f"font-size: 11px; color: {colors['text_muted']}; font-weight: bold;")
                else:
                    self.stats_label2.setStyleSheet("font-size: 11px; color: #81B4E0; font-weight: bold;")
        except Exception:
            pass

        # Repaint rows so OT/NC custom colors remain consistent after theme toggle.
        try:
            self.filter_cases()
        except Exception:
            pass
